#!/usr/bin/env python3
"""
syncer.py - hack.CCM Cloud Sync & Email Dispatch Engine
=========================================================
Handles git sync (with automatic backup tags), email dispatch, and
subscriber list sync — all via a single script with mode flags.

Default mode is git sync (`--mode all`). Email/subscribers require explicit flags.

USAGE (single commands or combine):
    python syncer.py                              # Git add -A, commit, push (default)
    python syncer.py --mode all                   # Git add -A, commit, push
    python syncer.py --mode data                  # Git all except main_app.py
    python syncer.py --mode web                   # Git main_app.py only
    python syncer.py --mode pearls                # Git pearls.json + sent_summaries.json
    python syncer.py --mode email                 # Dispatch pending emails
    python syncer.py --mode subscribers           # Sync Google Sheets -> emails.csv
    python syncer.py --mode full                  # subscribers -> email -> all (sequential)
    python syncer.py --mode all --verify          # Pre-flight health check before sync
    python syncer.py --dry-run                    # Preview only, no changes
    python syncer.py --verbose                      # Detailed logging
"""

import os
import sys
import re
import csv
import json
import time
import smtplib
import argparse
import subprocess
from datetime import datetime
from email.message import EmailMessage

from acumen_core.config import (
    PROJECT_DIR, EXCEL_TRACKER_FILE, JSON_TRACKER_FILE,
    OUTPUT_DIR, PEARLS_JSON,
)
from acumen_core.tracking import load_all_entries_from_json, save_json_atomic

# Fix stdout encoding
if sys.stdout.encoding != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except AttributeError:
        pass


# =====================================================================
# CONFIGURATION - Edit these for easy changes
# =====================================================================

# --- Git settings ---
GIT_REMOTE = "origin"
GIT_BRANCH = "main"
COMMIT_PREFIX = "hack.CCM Sync"
AUTO_COMMIT = True  # auto-commit if there are staged changes
BACKUP_BEFORE_PUSH = True  # create a local tag before pushing (backup-YYYYMMDD-HHMMSS)

# --- Vercel / Web App ---
WEB_APP_URL = "hack-ccm-acumen-microlearning.vercel.app"

# --- Google Forms URLs ---
FEEDBACK_FORM_URL = "https://docs.google.com/forms/d/e/1FAIpQLSd6xRmmimmVc0Sv4AeNls-oxLR6k_zX8D_QERFZwPP6zlfjRw/viewform?usp=header"
SUBSCRIBE_FORM_URL = "https://docs.google.com/forms/d/e/1FAIpQLSffsrF8DPWaTa-03XisMqSU5Da_8QdE-JrINdDP5iRmvWAI8Q/viewform?usp=header"
UNSUBSCRIBE_FORM_URL = "https://docs.google.com/forms/d/e/1FAIpQLScz864mkLh5AqBYVzAh573hWu98NdmwwPC2vaU1lfBE3WHHHg/viewform?usp=header"

# --- Subscriber sync (Google Sheets) ---
SUBSCRIBE_SHEET_ID = "1JJ5BtmhRPh_d3fmr433YHHZ9X6CjmvzGWa2-IrJwqYQ"
UNSUBSCRIBE_SHEET_ID = "1oD0oR7lW7mhDBRWuNPH5Z3fjtbwwZS8kHQT9qOj1dQA"
SUBSCRIBE_CSV_URL = f"https://docs.google.com/spreadsheets/d/{SUBSCRIBE_SHEET_ID}/export?format=csv"
UNSUBSCRIBE_CSV_URL = f"https://docs.google.com/spreadsheets/d/{UNSUBSCRIBE_SHEET_ID}/export?format=csv"
LOCAL_EMAILS_FILE = os.path.join(PROJECT_DIR, "emails.csv")

# --- Email SMTP (hardcoded credentials) ---
HARDCODED_SMTP_USER = "ronakjavia27@gmail.com"
HARDCODED_SMTP_PASS = "yioonbrhnnlhfhmz"
HARDCODED_SMTP_HOST = "smtp.gmail.com"
HARDCODED_SMTP_PORT = 587

# --- Email visual colors ---
COLOR_BG = "#FDFBF7"
COLOR_BTN = "#1D4ED8"
COLOR_BTN_TEXT = "#FFFFFF"
COLOR_TEXT = "#111827"
COLOR_BORDER = "#EFECE6"

# --- Email settings ---
EMAIL_SUBJECT_PREFIX = "hack.CCM Briefing"
EMAIL_INTERACTIVE = True  # prompt for selection or send all pending


# =====================================================================
# HELPER - Run git command
# =====================================================================
def run_git(*args, capture=True, check=False):
    """Run a git command, return CompletedProcess."""
    cmd = ["git"] + list(args)
    return subprocess.run(cmd, capture_output=capture, text=True, check=check, cwd=PROJECT_DIR)


def git_status_porcelain():
    """Return git status --porcelain output."""
    result = run_git("status", "--porcelain")
    return result.stdout.strip() if result.returncode == 0 else ""


def git_diff_cached_stat():
    """Return staged changes summary."""
    result = run_git("diff", "--cached", "--stat")
    return result.stdout.strip() if result.returncode == 0 else ""


# =====================================================================
# PRE-FLIGHT VALIDATION
# =====================================================================
def verify_before_sync(verbose=False):
    """Check sent_summaries.json consistency before pushing."""
    print("  [VERIFY] Pre-flight health check...")
    entries = load_all_entries_from_json()
    if not entries:
        print("  [!] WARNING: sent_summaries.json is empty or missing")
        return False

    # Check for duplicates
    file_names = [e.get("file_name", "") for e in entries]
    dupes = [f for f in file_names if file_names.count(f) > 1]
    if dupes:
        print(f"  [!] WARNING: {len(set(dupes))} duplicate file_names found")
        if verbose:
            for d in set(dupes):
                print(f"    - {d}")

    # Check web-approved count
    web_approved = sum(1 for e in entries if str(e.get("show_on_web", "")).lower() == "yes")
    print(f"  [OK] {len(entries)} entries loaded | {web_approved} web-approved")

    # Check pearls.json exists
    if os.path.exists(PEARLS_JSON):
        with open(PEARLS_JSON, "r", encoding="utf-8") as f:
            try:
                pearls = json.load(f)
                print(f"  [OK] pearls.json: {len(pearls)} entries")
            except Exception:
                print("  [!] WARNING: pearls.json is malformed")
    else:
        print("  [!] WARNING: pearls.json not found")

    return True


# =====================================================================
# GIT SYNC MODES
# =====================================================================
def git_sync(mode="all", verbose=False, skip_verify=False, dry_run=False):
    """Stage, commit, and push based on mode."""
    print(f"\n  [GIT-SYNC] Mode: {mode}")

    if not skip_verify:
        verify_before_sync(verbose)

    status = git_status_porcelain()
    if not status:
        print("  [OK] Workspace already clean — nothing to push.")
        return True

    print(f"  Pending changes detected:")
    if verbose:
        for line in status.split("\n"):
            print(f"    {line}")

    if dry_run:
        print("  [Dry-run] Would stage, commit, and push.")
        by_mode = {
            "web": ["main_app.py"],
            "data": ["all except main_app.py"],
            "pearls": ["pearls.json", "sent_summaries.json"],
            "all": ["all (respecting .gitignore)"],
        }
        for f in by_mode.get(mode, by_mode["all"]):
            print(f"    git add {f}")
        print(f"    git commit -m \"...\"")
        print(f"    git push {GIT_REMOTE} {GIT_BRANCH}")
        return True

    # Stage based on mode
    if mode == "web":
        print("  Staging main_app.py only...")
        run_git("add", "main_app.py")
    elif mode == "data":
        print("  Staging all files except main_app.py...")
        run_git("add", "-A")
        run_git("reset", "main_app.py")
    elif mode == "pearls":
        print("  Staging pearls.json + sent_summaries.json...")
        run_git("add", "pearls.json")
        run_git("add", "sent_summaries.json")
    else:  # all
        print("  Staging all changes (respecting .gitignore)...")
        run_git("add", "-A")

    staged = git_diff_cached_stat()
    if staged:
        print(f"  Staged:\n    {staged.replace(chr(10), chr(10) + '    ')}")
    else:
        print("  Nothing staged after filtering. Skipping commit.")
        return True

    # Commit
    if AUTO_COMMIT:
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        commit_msg = f"{COMMIT_PREFIX} - {now_str} ({mode})"
        result = run_git("commit", "-m", commit_msg)
        if result.returncode != 0:
            print(f"  [!] Commit failed: {result.stderr.strip()}")
            return False
        print(f"  Committed: {commit_msg}")
    else:
        print("  AUTO_COMMIT=False — staged but not committed.")
        return True

    # Backup (local tag before push)
    if BACKUP_BEFORE_PUSH:
        tag_name = f"backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        result = run_git("tag", tag_name)
        if result.returncode == 0:
            print(f"  [OK] Backup tag created: {tag_name}")
        else:
            print(f"  [!] Backup tag failed: {result.stderr.strip()}")

    # Push
    print(f"  Pushing to {GIT_REMOTE}/{GIT_BRANCH}...")
    result = run_git("push", GIT_REMOTE, GIT_BRANCH)
    if result.returncode == 0:
        print("  [OK] Push successful!")
        return True
    else:
        print(f"  [X] Push failed: {result.stderr.strip()}")
        return False


# =====================================================================
# EMAIL DISPATCH
# =====================================================================
def normalize_doi_url(raw_doi):
    """Sanitize raw DOIs into full links."""
    val = str(raw_doi).strip()
    if not val or val.lower() in ["none", "nan", ""]:
        return None
    if val.startswith("http://") or val.startswith("https://"):
        return val
    return f"https://doi.org/{val}"


def generate_button_ribbon(doi_url=None):
    """Generate HTML button row for emails."""
    app_btn = f'<a href="https://{WEB_APP_URL}" target="_blank" style="background-color:{COLOR_BTN}; color:{COLOR_BTN_TEXT}; padding:10px 16px; border-radius:6px; font-weight:bold; font-size:12px; text-decoration:none; display:inline-block; margin:4px;">Knowledge Portal</a>'
    doi_btn = ""
    if doi_url:
        doi_btn = f'<a href="{doi_url}" target="_blank" style="background-color:{COLOR_BTN}; color:{COLOR_BTN_TEXT}; padding:10px 16px; border-radius:6px; font-weight:bold; font-size:12px; text-decoration:none; display:inline-block; margin:4px;">Source Article</a>'
    sub_btn = f'<a href="{SUBSCRIBE_FORM_URL}" target="_blank" style="background-color:{COLOR_BTN}; color:{COLOR_BTN_TEXT}; padding:10px 16px; border-radius:6px; font-weight:bold; font-size:12px; text-decoration:none; display:inline-block; margin:4px;">Subscribe</a>'
    return f"""
    <div style="text-align: center; margin-top: 15px; margin-bottom: 15px; padding: 10px; background-color: {COLOR_BORDER}; border-radius: 8px;">
        {app_btn} {doi_btn} {sub_btn}
    </div>
    """


def convert_markdown_to_html(markdown_text, doi_url=None):
    """Convert markdown summary to HTML email body."""
    lines = markdown_text.split("\n")
    html_lines = [generate_button_ribbon(doi_url)]
    in_list = False

    for line in lines:
        line_str = line.strip()
        if not line_str:
            if in_list:
                html_lines.append("</ul>")
                in_list = False
            continue

        line_str = re.sub(r'\*\*(.*?)\*\*', r'<strong>\1</strong>', line_str)

        if line_str.startswith("###"):
            if in_list: html_lines.append("</ul>"); in_list = False
            html_lines.append(f'<h3 style="color:#000000; border-bottom:1px solid {COLOR_BORDER}; padding-bottom:4px; font-size:16px; margin-top:20px; margin-bottom:8px;">{line_str[3:].strip()}</h3>')
        elif line_str.startswith("##"):
            if in_list: html_lines.append("</ul>"); in_list = False
            html_lines.append(f'<h2 style="color:#000000; font-size:20px; margin-top:22px; margin-bottom:10px;">{line_str[2:].strip()}</h2>')
        elif line_str.startswith("-") or line_str.startswith("*"):
            if not in_list:
                html_lines.append('<ul style="margin-left:20px; padding-left:0; margin-bottom:12px; list-style-type:disc;">')
                in_list = True
            html_lines.append(f'<li style="margin-bottom:6px; line-height:1.5;">{line_str[1:].strip()}</li>')
        else:
            if in_list: html_lines.append("</ul>"); in_list = False
            html_lines.append(f'<p style="margin-bottom:12px; line-height:1.6; text-align:justify;">{line_str}</p>')

    if in_list:
        html_lines.append("</ul>")

    html_lines.append(generate_button_ribbon(doi_url))
    html_lines.append(f"""
        <hr style="border:0; border-top:1px solid {COLOR_BORDER}; margin-top:25px; margin-bottom:15px;" />
        <div style="text-align:center; font-size:11px; color:#4B5563;">
            This dispatch was sent directly to your registered address.
            Want to share feedback? <a href="{FEEDBACK_FORM_URL}" target="_blank" style="color:{COLOR_BTN};">Click Here</a>.<br/>
            To opt-out from future clinical briefs instantly, you can <a href="{UNSUBSCRIBE_FORM_URL}" target="_blank" style="color:{COLOR_BTN}; text-decoration:underline;">Unsubscribe here</a> safely.
        </div>
    """)

    return f"""
    <div style="background-color:{COLOR_BG}; padding:20px; font-family:'Georgia', serif; color:{COLOR_TEXT};">
        <div style="max-width:650px; margin:0 auto; background-color:#FFFFFF; border:1px solid {COLOR_BORDER}; border-radius:12px; padding:30px; box-shadow:0 2px 8px rgba(0,0,0,0.01);">
            {"".join(html_lines)}
        </div>
    </div>
    """


def send_email(subject, html_content, recipient):
    """Send one email via SMTP."""
    if not HARDCODED_SMTP_USER or not HARDCODED_SMTP_PASS or "@" not in HARDCODED_SMTP_USER:
        print("  [X] SMTP credentials missing")
        return False

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = f"{EMAIL_SUBJECT_PREFIX} <{HARDCODED_SMTP_USER}>"
    msg["To"] = recipient
    msg.set_content("Please switch to an HTML-supported client to view this brief.")
    msg.add_alternative(html_content, subtype="html")

    try:
        with smtplib.SMTP(HARDCODED_SMTP_HOST, HARDCODED_SMTP_PORT, timeout=15) as server:
            server.starttls()
            server.login(HARDCODED_SMTP_USER, HARDCODED_SMTP_PASS)
            server.send_message(msg)
        return True
    except Exception as e:
        print(f"  [X] SMTP failed for {recipient}: {e}")
        return False


def dispatch_emails(dry_run=False, verbose=False, send_all=False):
    """Dispatch pending emails (where email_pushed != 'Yes')."""
    print("\n  [EMAIL] Dispatching pending articles...")

    if not os.path.exists(EXCEL_TRACKER_FILE):
        print(f"  [X] Excel tracker not found: {EXCEL_TRACKER_FILE}")
        return 0

    try:
        from openpyxl import load_workbook
    except Exception:
        print("  [X] openpyxl not installed")
        return 0

    # Load subscribers
    subscribers = []
    if os.path.exists(LOCAL_EMAILS_FILE):
        with open(LOCAL_EMAILS_FILE, mode="r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                if row and len(row) > 0:
                    email_val = row[0].strip()
                    if email_val and "@" in email_val:
                        subscribers.append(email_val)

    if not subscribers:
        print(f"  [!] No subscribers in {os.path.basename(LOCAL_EMAILS_FILE)}")
        if not dry_run:
            return 0

    print(f"  Loaded {len(subscribers)} subscribers")

    # Scan Excel for pending articles
    wb = load_workbook(EXCEL_TRACKER_FILE)
    ws = wb["Registry Logs"]

    pending_articles = []
    for row in range(2, ws.max_row + 1):
        already_sent = str(ws.cell(row=row, column=10).value).strip().lower()  # Column J
        if already_sent != "yes":
            file_name = ws.cell(row=row, column=2).value  # Column B
            raw_doi = ws.cell(row=row, column=6).value  # Column F
            if not file_name:
                continue

            # Find the JSON file on disk
            base_json_name = os.path.splitext(str(file_name))[0] + ".json"
            full_json_path = None
            for root, dirs, files in os.walk(OUTPUT_DIR):
                if base_json_name in files:
                    full_json_path = os.path.join(root, base_json_name)
                    break

            if full_json_path and os.path.exists(full_json_path):
                try:
                    with open(full_json_path, "r", encoding="utf-8") as jf:
                        payload = json.load(jf)
                    paper_title = payload.get("title", payload.get("paper_name", file_name))
                    markdown = payload.get("clinical_summary_markdown", "")
                    if not markdown:
                        # Build from sections
                        md_parts = []
                        for s in payload.get("sections", []):
                            md_parts.append(f"## {s.get('heading', '')}\n{s.get('content', '')}")
                        markdown = "\n\n".join(md_parts)
                    if markdown:
                        pending_articles.append({
                            "row_idx": row,
                            "title": paper_title,
                            "doi": normalize_doi_url(raw_doi),
                            "markdown": markdown,
                        })
                except Exception as e:
                    if verbose:
                        print(f"  [skip] Failed to load {base_json_name}: {e}")

    if not pending_articles:
        print("  No pending unsent articles detected.")
        return 0

    print(f"\n  Pending unsent articles: {len(pending_articles)}")
    for idx, item in enumerate(pending_articles, start=1):
        print(f"    [{idx}] Row {item['row_idx']} -> {item['title'][:70]}")

    if dry_run:
        print("\n  [Dry-run] No emails sent.")
        return len(pending_articles)

    # Selection
    target_items = pending_articles
    if EMAIL_INTERACTIVE and not send_all and len(pending_articles) > 1:
        print("\n  Select dispatch:")
        print("    Type number(s) e.g. 1, 3")
        print("    Type 'all' to send everything")
        user_selection = input("\n  Selection: ").strip().lower()
        if user_selection == "all":
            target_items = pending_articles
        else:
            try:
                indices = [int(x.strip()) - 1 for x in user_selection.split(",") if x.strip().isdigit()]
                target_items = [pending_articles[i] for i in indices if 0 <= i < len(pending_articles)]
            except Exception:
                print("  [X] Invalid selection.")
                return 0

    if not target_items:
        print("  No valid articles selected.")
        return 0

    # Dispatch
    dispatched = 0
    for item in target_items:
        print(f"\n  Sending: {item['title'][:70]}")
        html = convert_markdown_to_html(item["markdown"], item["doi"])
        subject = f"{EMAIL_SUBJECT_PREFIX}: {item['title']}"

        success_all = True
        for email in subscribers:
            if not send_email(subject, html, email):
                success_all = False
                break

        if success_all:
            row = item["row_idx"]
            ws.cell(row=row, column=10).value = "Yes"       # Email Pushed
            ws.cell(row=row, column=14).value = "Yes"       # show_on_web
            ws.cell(row=row, column=12).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            dispatched += 1
            print(f"    [OK] Sent to {len(subscribers)} subscribers")

    if dispatched > 0:
        temp_file = EXCEL_TRACKER_FILE + ".tmp"
        wb.save(temp_file)
        os.replace(temp_file, EXCEL_TRACKER_FILE)

        # Also update JSON tracker
        entries = load_all_entries_from_json()
        sent_titles = {item["title"] for item in target_items if item in target_items[:dispatched]}
        for entry in entries:
            title = entry.get("title", "")
            if title in sent_titles:
                entry["email_pushed"] = "Yes"
                entry["email_pushed_date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                entry["show_on_web"] = "Yes"
        save_json_atomic(JSON_TRACKER_FILE, entries)

        print(f"\n  [OK] Dispatched {dispatched}/{len(target_items)} articles.")
        print(f"  Excel + JSON trackers updated.")
    else:
        print("\n  [!] No articles were successfully dispatched.")

    return dispatched


# =====================================================================
# SUBSCRIBER SYNC (Google Sheets -> emails.csv)
# =====================================================================
def sync_subscribers(dry_run=False, verbose=False):
    """Fetch subscribe/unsubscribe sheets, merge chronologically, write emails.csv."""
    print("\n  [SUBSCRIBERS] Syncing from Google Sheets...")

    try:
        import requests
    except Exception:
        print("  [X] requests not installed")
        return 0

    lifecycle_events = []

    # 1. Subscribe sheet
    print(f"  Fetching subscribe sheet...")
    try:
        response = requests.get(SUBSCRIBE_CSV_URL, timeout=10)
        if response.status_code == 200:
            lines = response.content.decode("utf-8").splitlines()
            reader = csv.reader(lines)
            next(reader, None)  # skip header
            for row in reader:
                if row and len(row) > 2:
                    timestamp = row[0].strip()
                    email = row[2].strip().lower()
                    if "@" in email and timestamp:
                        lifecycle_events.append((timestamp, email, "SUBSCRIBE"))
        else:
            print(f"  [!] Subscribe fetch failed: HTTP {response.status_code}")
    except Exception as e:
        print(f"  [!] Subscribe fetch error: {e}")

    # 2. Unsubscribe sheet
    print(f"  Fetching unsubscribe sheet...")
    try:
        response = requests.get(UNSUBSCRIBE_CSV_URL, timeout=10)
        if response.status_code == 200:
            lines = response.content.decode("utf-8").splitlines()
            reader = csv.reader(lines)
            next(reader, None)
            for row in reader:
                if row and len(row) > 1:
                    timestamp = row[0].strip()
                    email = row[1].strip().lower()
                    if "@" in email and timestamp:
                        lifecycle_events.append((timestamp, email, "UNSUBSCRIBE"))
        else:
            print(f"  [!] Unsubscribe fetch failed: HTTP {response.status_code}")
    except Exception as e:
        print(f"  [!] Unsubscribe fetch error: {e}")

    if not lifecycle_events:
        print("  No transactions detected. Local list unchanged.")
        return 0

    # 3. Sort chronologically
    lifecycle_events.sort(key=lambda x: x[0])

    # 4. Apply state changes
    final = set()
    for timestamp, email, action in lifecycle_events:
        if action == "SUBSCRIBE":
            final.add(email)
        elif action == "UNSUBSCRIBE":
            final.discard(email)

    print(f"  Subscribe events:  {sum(1 for _, _, a in lifecycle_events if a == 'SUBSCRIBE')}")
    print(f"  Unsubscribe events: {sum(1 for _, _, a in lifecycle_events if a == 'UNSUBSCRIBE')}")
    print(f"  Final active:       {len(final)}")

    if dry_run:
        print("  [Dry-run] emails.csv not updated.")
        return len(final)

    with open(LOCAL_EMAILS_FILE, mode="w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for email in sorted(final):
            writer.writerow([email])

    print(f"  [OK] Synced to {os.path.basename(LOCAL_EMAILS_FILE)}")
    return len(final)


# =====================================================================
# FULL SYNC SEQUENCE
# =====================================================================
def full_sync(dry_run=False, verbose=False):
    """Sequential: subscribers -> email -> git all."""
    print("\n  [FULL] Running full sync sequence...")
    print("  " + "=" * 50)

    print("\n  Step 1/3: Subscriber sync")
    sync_subscribers(dry_run=dry_run, verbose=verbose)

    print("\n  Step 2/3: Email dispatch")
    if EMAIL_INTERACTIVE and not dry_run:
        dispatch_emails(dry_run=dry_run, verbose=verbose, send_all=True)
    else:
        dispatch_emails(dry_run=dry_run, verbose=verbose, send_all=True)

    print("\n  Step 3/3: Git sync (all)")
    if not dry_run:
        git_sync(mode="all", verbose=verbose, skip_verify=False, dry_run=False)

    print("\n  " + "=" * 50)
    print("  [FULL] Complete!\n")


# =====================================================================
# MAIN ENTRY POINT
# =====================================================================
def main():
    parser = argparse.ArgumentParser(
        description="hack.CCM Cloud Sync & Email Dispatch Engine",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python syncer.py --mode all                   # Git add -A, commit, push
  python syncer.py --mode data                  # Git all except main_app.py
  python syncer.py --mode web                   # Git main_app.py only
  python syncer.py --mode pearls                # Git pearls.json + sent_summaries.json
  python syncer.py --mode email                  # Dispatch pending emails
  python syncer.py --mode subscribers            # Sync Google Sheets -> emails.csv
  python syncer.py --mode full                   # subscribers -> email -> all
  python syncer.py --mode all --verify           # Pre-flight check before sync
  python syncer.py --dry-run                     # Preview only
  python syncer.py --verbose                       # Detailed logging
        """,
    )
    parser.add_argument("--mode", choices=["all", "data", "web", "pearls", "email", "subscribers", "full"],
                        default="all", help="What to sync")
    parser.add_argument("--verify", action="store_true", help="Pre-flight health check before sync")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, no changes")
    parser.add_argument("--verbose", action="store_true", help="Detailed logging")
    args = parser.parse_args()

    if args.mode == "full":
        full_sync(dry_run=args.dry_run, verbose=args.verbose)
    elif args.mode == "email":
        dispatch_emails(dry_run=args.dry_run, verbose=args.verbose, send_all=not EMAIL_INTERACTIVE)
    elif args.mode == "subscribers":
        sync_subscribers(dry_run=args.dry_run, verbose=args.verbose)
    else:
        # Git modes: all, data, web, pearls
        git_sync(mode=args.mode, verbose=args.verbose, skip_verify=not args.verify, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
