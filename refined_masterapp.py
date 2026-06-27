import os
import sys
import time
import json
import re
from datetime import datetime
from pypdf import PdfReader
from google import genai
from google.genai import types

# 1. IMPORT AND RUN DOTENV FIRST
from dotenv import load_dotenv

# Dynamically resolve absolute path to eliminate any terminal folder bugs
script_dir = os.path.dirname(os.path.abspath(__file__))
env_path = os.path.join(script_dir, '.env')
load_dotenv(dotenv_path=env_path)

# Third-Party Excel Library
import openpyxl
from openpyxl import Workbook, load_workbook

# =====================================================================
# ⚙️ CENTRAL CONTROL DECK & CONFIGURATIONS
# =====================================================================
BASE_INPUT_DIR = "./input_pdfs"
SUB_DIRS = {
    "articles": os.path.join(BASE_INPUT_DIR, "articles"),
    "guidelines": os.path.join(BASE_INPUT_DIR, "guidelines"),
    "other": os.path.join(BASE_INPUT_DIR, "other")
}

OUTPUT_DIR = "./output_files"
EXCEL_TRACKER_FILE = "./sent_summaries.xlsx"
SPECIALTIES_FILE = "./specialties.txt"
ARTICLE_TYPES_FILE = "./article_types.txt"

# Model Mapping based on specifications
MODEL_ARTICLES = "gemini-3.5-flash"
MODEL_GUIDELINES = "gemini-3.5-flash"  
MODEL_BACKUP = "gemini-2.5-pro"

PRIMARY_API_KEY = os.getenv("PRIMARY_GEMINI_API_KEY")
BACKUP_API_KEY = os.getenv("BACKUP_GEMINI_API_KEY")

if not PRIMARY_API_KEY and not BACKUP_API_KEY:
    print("🚨 CRITICAL ERROR: No API keys found! Please check your .env file layout.")
    sys.exit(1)

PRIMARY_CLIENT = genai.Client(api_key=PRIMARY_API_KEY) if PRIMARY_API_KEY else None
BACKUP_CLIENT = genai.Client(api_key=BACKUP_API_KEY) if BACKUP_API_KEY else None

# =====================================================================
# 📋 CONTROLLED VOCABULARY MATRICES MANAGEMENT
# =====================================================================
def load_allowed_vocabulary(file_path, default_list):
    """Loads a validation text listing or populates defaults if missing."""
    if not os.path.exists(file_path):
        with open(file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(default_list))
        return default_list
    with open(file_path, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]

def validate_and_scrub_metadata(extracted_val, allowed_list):
    """Programmatically sanitizes values to catch fuzzy matches or default to Other."""
    cleaned_input = str(extracted_val).strip().lower()
    
    # Exact Match check
    for allowed_term in allowed_list:
        if cleaned_input == allowed_term.strip().lower():
            return allowed_term
            
    # Fuzzy / Slight partial match check
    for allowed_term in allowed_list:
        clean_allowed = allowed_term.strip().lower()
        if clean_allowed in cleaned_input or cleaned_input in clean_allowed:
            if clean_allowed not in ["other", "general", "unclassified"]:
                return allowed_term
                
    # Fallback to Other if no alignment matches can be mapped
    for allowed_term in allowed_list:
        if allowed_term.strip().lower() == "other":
            return allowed_term
    return "Other"

# Initialize Vocabulary Filters
DEFAULT_SPECIALTIES = ["Critical Care Medicine", "Cardiovascular", "Neurology", "Nephrology", "Pulmonology", "Other"]
DEFAULT_TYPES = ["Guideline", "Review", "Meta-Analysis", "Trial", "Other"]

ALLOWED_SPECIALTIES = load_allowed_vocabulary(SPECIALTIES_FILE, DEFAULT_SPECIALTIES)
ALLOWED_TYPES = load_allowed_vocabulary(ARTICLE_TYPES_FILE, DEFAULT_TYPES)

# =====================================================================
# 📊 EXCEL TRANSACTION TRACKING MATRIX
# =====================================================================
EXCEL_HEADERS = [
    "Serial Number", "File Name", "Paper/Guideline Name", "Primary Authors", 
    "Journal Name", "DOI", "System", "Type of Article", "MD Generated", 
    "Email Pushed", "Summary Saved Date", "Email Pushed Date", "Parsing Notes", "show_on_web"
]

def initialize_system_paths():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    for folder_path in SUB_DIRS.values():
        os.makedirs(folder_path, exist_ok=True)
                
    if not os.path.exists(EXCEL_TRACKER_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registry Logs"
        ws.append(EXCEL_HEADERS)
        wb.save(EXCEL_TRACKER_FILE)
        print(f"📊 Initialized ledger registry tracker at: {EXCEL_TRACKER_FILE}")

def load_processed_files_from_excel():
    if not os.path.exists(EXCEL_TRACKER_FILE):
        return set()
    try:
        wb = load_workbook(EXCEL_TRACKER_FILE, read_only=True)
        ws = wb["Registry Logs"]
        processed = set()
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if len(row) > 1 and row[1]: 
                processed.add(str(row[1]).strip())
        return processed
    except Exception:
        return set()

def log_transaction_to_excel(file_name, metadata, parsing_notes="Success"):
    retries = 5
    for attempt in range(retries):
        try:
            wb = load_workbook(EXCEL_TRACKER_FILE)
            ws = wb["Registry Logs"]
            next_serial = ws.max_row
            
            row_data = [
                next_serial, file_name,
                metadata.get("paper_name", "Unknown Title"),
                metadata.get("primary_authors", "Unknown Authors"),
                metadata.get("journal_name", "Unknown Journal"),
                metadata.get("doi", "None"),
                metadata.get("system", "Other"),
                metadata.get("type_of_article", "Other"),
                "Yes", "No",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "",
                parsing_notes, "No"
            ]
            ws.append(row_data)
            wb.save(EXCEL_TRACKER_FILE)
            return
        except PermissionError:
            time.sleep(1.5)

# =====================================================================
# 🧠 FAIL-SAFE MODEL EXECUTION ENGINE
# =====================================================================
def execute_gemini_with_fallback(prompt, system_instruction, target_model, response_json=False):
    execution_matrix = [
        (PRIMARY_CLIENT, target_model),
        (BACKUP_CLIENT, target_model),
        (PRIMARY_CLIENT, MODEL_BACKUP)
    ]
    
    config = types.GenerateContentConfig(
        system_instruction=system_instruction,
        temperature=0.1,
        response_mime_type="application/json" if response_json else "text/plain"
    )
    
    last_error = None
    for client, model in execution_matrix:
        if not client:
            continue
        for attempt in range(3):
            try:
                response = client.models.generate_content(
                    model=model,
                    contents=prompt,
                    config=config
                )
                if response_json:
                    return json.loads(response.text.strip())
                return response.text.strip()
            except Exception as e:
                last_error = e
                time.sleep(3)
                
    raise RuntimeError(f"🚨 Critical Failure: API key layers exhausted. Error: {last_error}")

def generate_final_structured_payload(uploaded_file, category_tag):
    """Analyzes the file object natively via Google Files API and enforces vocabulary matching rules."""
    target_model = MODEL_GUIDELINES if category_tag == "guidelines" else MODEL_ARTICLES
    
    system_instruction = f"""You are an elite, world-class Clinical Research Ingestion Engine specializing in Critical Care Medicine and complex medical subspecialties. Your purpose is to digest medical literature (Guidelines, Clinical Trials, Meta-Analyses, or narrative Review articles) with exhaustive, granular precision.

You MUST respond strictly with a single valid JSON object matching this exact schema:
{{
  "paper_name": "Official Title of the Paper/Guideline/Review",
  "primary_authors": "Lead Author et al.",
  "journal_name": "Official Publishing Medical Journal Name",
  "doi": "Clean DOI identifier string",
  "system": "Choose exactly one: {', '.join(ALLOWED_SPECIALTIES)}",
  "type_of_article": "Choose exactly one: {', '.join(ALLOWED_TYPES)}",
  "clinical_summary_markdown": "Your detailed, un-truncated evaluation goes here"
}}

CRITICAL EXTRACTION DIRECTIVES FOR THE 'clinical_summary_markdown' FIELD:
1. Target an exhaustive length of 2,500 to 4,000 words for the summary body text. Never summarize or omit dense textual sub-layers.
2. Read and step into every sequential subsection of the PDF. Do not bypass paragraphs or state generalized conclusions.
3. DO NOT use LaTeX formatting or math syntax. Write out clinical parameters in clean plain-text notation (e.g., write 'PaO2/FiO2 ratio <= 150 mmHg' instead of math blocks).

You MUST strictly follow this exact structural markdown layout template inside your JSON value string:

## 📋 CORE EXECUTIVE OVERVIEW & CLINICAL RATIONALE
[Provide a dense, expert-level high-yield summary of the baseline clinical problems, current controversies, or gaps in knowledge that this document addresses.]

## 🩺 CONDITIONAL SECTION: METICULOUS TRIAL METHODOLOGY & DESIGN
*(NOTE: If this document is a comprehensive narrative Review article or a Guideline rather than an isolated Trial, adapt this section to outline the search strategy, meta-regression rules, or systemic consensus frameworks used.)*
- **Study Framework:** [Phase, multi-center scale, blinding status, randomization layout, or guidelines panel consensus development profile]
- **Inclusion / Selection Criteria:** [Exhaustive bedside parameters, comorbidity metrics, diagnostic parameters, and absolute scoring gates]
- **Exclusion / Elimination Criteria:** [List every exact organ failure threshold, time window restriction, or safety exclusion boundary]
- **Management Protocol & Infusion Architecture:** [Detail exact treatment arms, drug generic names, specific dosing rules, extended/intermittent infusion parameters, titration rules, and control arm settings explicitly]

## 🔬 SEQUENTIAL SUBSECTION ANALYSIS & CORE THEMATIC ARCHITECTURE
*(CRITICAL FOR REVIEWS AND GUIDELINES: Traverse the document section-by-section. Break down every major physiological system, clinical sub-topic, disease state, diagnostic algorithm, or section heading discussed in the paper.)*
### [Heading / Core Topic 1 from Paper]
- **Detailed Synthesis:** [Write an information-dense, textbook-grade evaluation of the underlying mechanisms, diagnostic pathways, clinical criteria, or drug options detailed in this section.]
- **Granular Data Points:** [Extract every exact number, threshold, drug dose, or biomarker parameter mentioned natively.]

### [Heading / Core Topic 2 from Paper]
- **Detailed Synthesis:** [Continue step-by-step through the next sequential topic/section of the article with identical granular detail.]
- **Granular Data Points:** [Extract all associated numerical metrics, physiological endpoints, or expert classifications.]

### [Heading / Core Topic 3 from Paper (Add more as required by the document)]
- **Detailed Synthesis:** [Expand comprehensively on subsequent sub-topics, guidelines tiers, or disease updates.]
- **Granular Data Points:** [Maintain raw ground truth without generalized statements.]

## 📊 STATISTICAL RESULTS, ENDPOINTS & SAFETY OUTCOMES
- **Primary Metrics / Key Recommendations:** [List primary outcomes, trial findings, or major grade-A guideline recommendations with accompanying Hazard Ratios (HR), Confidence Intervals (95% CI), and absolute p-values if available.]
- **Secondary Parameters & Subgroup Findings:** [Detail secondary physiological goals, subgroup analysis like chronic kidney disease cohorts, geriatric variables, or secondary endpoints.]
- **Adverse Events & Safety Disclosures:** [List all complication incidence rates, toxicity thresholds, or reasons for premature treatment cessation.]

## 🧠 BEDSIDE CLINICAL APPLICATION & ICU BLUEPRINT MANAGEMENT
- **Methodological Strengths & Robustness:** [Meticulous assessment of evidence quality, bias mitigation, or structural trial/review completeness.]
- **Limitations, Blind Spots & Knowledge Gaps:** [Identify design vulnerabilities, potential confounding loops, industry funding footprints, or areas needing trial exploration.]
- **Bedside Protocol Blueprint & Titration Workflows:** [Provide a highly detailed, step-by-step practical implementation guide for immediate clinical use. Detail the exact clinical scenario, specific drug adjustment rules, targets, lab tracking markers, and explicit stopping rules based on active patient response metrics.]
"""

    prompt_content = [
        uploaded_file,
        "Execute a thorough, comprehensive section-by-section ingestion pass of the entire document. Ensure narrative review themes, consensus tables, and guideline tiers are extracted to the maximum structural depth."
    ]

    print(f"  🧠 Analyzing payload architecture via {target_model}...")
    return execute_gemini_with_fallback(prompt_content, system_instruction, target_model, response_json=True)

# =====================================================================
# 📥 EXTRACTION PROCESSING LAYER
# =====================================================================
def process_single_pdf(file_path, category, processed_history):
    file_name = os.path.basename(file_path)
    if file_name in processed_history:
        return

    print(f"\n⚡ Ingesting target medical asset [{category.upper()}]: {file_name}")
    uploaded_file = None
    client_to_use = PRIMARY_CLIENT if PRIMARY_CLIENT else BACKUP_CLIENT
    
    try:
        print("  📤 Streaming document to Google Files API Hub...")
        uploaded_file = client_to_use.files.upload(file=file_path)
        
        while uploaded_file.state.name == "PROCESSING":
            time.sleep(2)
            uploaded_file = client_to_use.files.get(name=uploaded_file.name)
            
        if uploaded_file.state.name == "FAILED":
            raise ValueError("Google Files API dropped document serialization layers.")

        structured_payload = generate_final_structured_payload(uploaded_file, category)
        
        # Programmatic scrubbing gates using the controlled lists
        scrubbed_specialty = validate_and_scrub_metadata(structured_payload.get("system", "Other"), ALLOWED_SPECIALTIES)
        scrubbed_type = validate_and_scrub_metadata(structured_payload.get("type_of_article", "Other"), ALLOWED_TYPES)
        
        # Override values inside payload to match exactly
        structured_payload["system"] = scrubbed_specialty
        structured_payload["type_of_article"] = scrubbed_type
        
        # Dynamic Storage Path Generation
        clean_system = "".join(x for x in scrubbed_specialty if x.isalnum() or x in "._- ").strip()
        clean_type = "".join(x for x in scrubbed_type if x.isalnum() or x in "._- ").strip()
        
        sharded_output_dir = os.path.join(OUTPUT_DIR, clean_system, clean_type)
        os.makedirs(sharded_output_dir, exist_ok=True)
        
        base_name = os.path.splitext(file_name)[0]
        destination_json_path = os.path.join(sharded_output_dir, f"{base_name}.json")
        
        with open(destination_json_path, "w", encoding="utf-8") as jf:
            json.dump(structured_payload, jf, indent=2, ensure_ascii=False)
        print(f"  🌐 Ingestion matrix sharded to filesystem: {destination_json_path}")
        
        log_transaction_to_excel(file_name, structured_payload, "Success")
        processed_history.add(file_name)
        
        if os.path.exists(file_path):
            os.remove(file_path)

    except Exception as process_error:
        print(f"  ❌ Extraction Failure targeting ({file_name}): {process_error}")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            with open("./error_logs.txt", "a", encoding="utf-8") as ef:
                ef.write(f"[{timestamp}] File: {file_name} | Error: {str(process_error)}\n")
        except Exception:
            pass

    finally:
        if uploaded_file:
            try:
                client_to_use.files.delete(name=uploaded_file.name)
            except Exception:
                pass

# =====================================================================
# 📡 RUNTIME WATCHER LOOP POLLING
# =====================================================================
if __name__ == "__main__":
    initialize_system_paths()
    history_log = load_processed_files_from_excel()
    print("🚀 Advanced Medical Ingestion Engine Active...")
    print(f"  🩺 Specialties List: {len(ALLOWED_SPECIALTIES)} fields loaded.")
    print(f"  📑 Article Formats: {len(ALLOWED_TYPES)} variants loaded.")
    
    try:
        loop_counter = 0
        while True:
            loop_counter += 1
            history_log = load_processed_files_from_excel()
            
            if loop_counter % 12 == 1:
                print(f"⏱️ [Heartbeat] Monitoring watch folders... Time: {datetime.now().strftime('%H:%M:%S')}")
            
            for category, folder_path in SUB_DIRS.items():
                if os.path.exists(folder_path):
                    pdf_files = [f for f in os.listdir(folder_path) if f.endswith('.pdf')]
                    
                    for file_name in pdf_files:
                        path = os.path.join(folder_path, file_name)
                        if file_name in history_log:
                            continue
                            
                        try:
                            initial_size = os.path.getsize(path)
                            time.sleep(1.5)
                            if os.path.getsize(path) == initial_size:
                                process_single_pdf(path, category, history_log)
                        except Exception:
                            continue
            time.sleep(5)
    except KeyboardInterrupt:
        print("\n🛑 Ingestion loop cleanly terminated.")