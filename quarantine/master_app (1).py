import os
import sys
import time
import json
import re
from datetime import datetime
from pypdf import PdfReader
from google import genai
from google.genai import types

# 1. IMPORT AND RUN DOTENV FIRST
from dotenv import load_dotenv

# Dynamically resolve absolute path to eliminate any terminal folder bugs
script_dir = os.path.dirname(os.path.abspath(__file__))
env_path = os.path.join(script_dir, '.env')
load_dotenv(dotenv_path=env_path)

# Third-Party Excel Library
import openpyxl
from openpyxl import Workbook, load_workbook

# =====================================================================
# ⚙️ CENTRAL CONTROL DECK & CONFIGURATIONS
# =====================================================================
BASE_INPUT_DIR = "./input_pdfs"
SUB_DIRS = {
    "articles": os.path.join(BASE_INPUT_DIR, "articles"),
    "guidelines": os.path.join(BASE_INPUT_DIR, "guidelines"),
    "other": os.path.join(BASE_INPUT_DIR, "other")
}

OUTPUT_DIR = "./output_files"
EXCEL_TRACKER_FILE = "./sent_summaries.xlsx"
SPECIALTIES_FILE = "./specialties.txt"
ARTICLE_TYPES_FILE = "./article_types.txt"

# Model Mapping based on specifications
MODEL_ARTICLES = "gemini-3.5-flash"
MODEL_GUIDELINES = "gemini-3.1-pro-preview"  
MODEL_BACKUP = "gemini-2.1-pro"

PRIMARY_API_KEY = os.getenv("PRIMARY_GEMINI_API_KEY")
BACKUP_API_KEY = os.getenv("BACKUP_GEMINI_API_KEY")

if not PRIMARY_API_KEY and not BACKUP_API_KEY:
    print("🚨 CRITICAL ERROR: No API keys found! Please check your .env file layout.")
    sys.exit(1)

PRIMARY_CLIENT = genai.Client(api_key=PRIMARY_API_KEY) if PRIMARY_API_KEY else None
BACKUP_CLIENT = genai.Client(api_key=BACKUP_API_KEY) if BACKUP_API_KEY else None

# =====================================================================
# 📋 CONTROLLED VOCABULARY MATRICES MANAGEMENT
# =====================================================================
def load_allowed_vocabulary(file_path, default_list):
    """Loads a validation text listing or populates defaults if missing."""
    if not os.path.exists(file_path):
        with open(file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(default_list))
        return default_list
    with open(file_path, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]

def validate_and_scrub_metadata(extracted_val, allowed_list):
    """Programmatically sanitizes values to catch fuzzy matches or default to Other."""
    cleaned_input = str(extracted_val).strip().lower()
    
    # Exact Match check
    for allowed_term in allowed_list:
        if cleaned_input == allowed_term.strip().lower():
            return allowed_term
            
    # Fuzzy / Slight partial match check
    for allowed_term in allowed_list:
        clean_allowed = allowed_term.strip().lower()
        if clean_allowed in cleaned_input or cleaned_input in clean_allowed:
            if clean_allowed not in ["other", "general", "unclassified"]:
                return allowed_term
                
    # Fallback to Other if no alignment matches can be mapped
    for allowed_term in allowed_list:
        if allowed_term.strip().lower() == "other":
            return allowed_term
    return "Other"

# Initialize Vocabulary Filters
DEFAULT_SPECIALTIES = ["Critical Care Medicine", "Cardiovascular", "Neurology", "Nephrology", "Pulmonology", "Other"]
DEFAULT_TYPES = ["Guideline", "Review", "Meta-Analysis", "Trial", "Other"]

ALLOWED_SPECIALTIES = load_allowed_vocabulary(SPECIALTIES_FILE, DEFAULT_SPECIALTIES)
ALLOWED_TYPES = load_allowed_vocabulary(ARTICLE_TYPES_FILE, DEFAULT_TYPES)

# =====================================================================
# 📊 EXCEL TRANSACTION TRACKING MATRIX
# =====================================================================
EXCEL_HEADERS = [
    "Serial Number", "File Name", "Paper/Guideline Name", "Primary Authors", 
    "Journal Name", "DOI", "System", "Type of Article", "MD Generated", 
    "Email Pushed", "Summary Saved Date", "Email Pushed Date", "Parsing Notes", "show_on_web"
]

def initialize_system_paths():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    for folder_path in SUB_DIRS.values():
        os.makedirs(folder_path, exist_ok=True)
                
    if not os.path.exists(EXCEL_TRACKER_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registry Logs"
        ws.append(EXCEL_HEADERS)
        wb.save(EXCEL_TRACKER_FILE)
        print(f"📊 Initialized ledger registry tracker at: {EXCEL_TRACKER_FILE}")

def load_processed_files_from_excel():
    if not os.path.exists(EXCEL_TRACKER_FILE):
        return set()
    try:
        wb = load_workbook(EXCEL_TRACKER_FILE, read_only=True)
        ws = wb["Registry Logs"]
        processed = set()
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if len(row) > 1 and row[1]: 
                processed.add(str(row[1]).strip())
        return processed
    except Exception:
        return set()

def log_transaction_to_excel(file_name, metadata, parsing_notes="Success"):
    retries = 5
    for attempt in range(retries):
        try:
            wb = load_workbook(EXCEL_TRACKER_FILE)
            ws = wb["Registry Logs"]
            next_serial = ws.max_row
            
            row_data = [
                next_serial, file_name,
                metadata.get("title", metadata.get("paper_name", "Unknown Title")),
                metadata.get("authors", metadata.get("primary_authors", "Unknown Authors")),
                metadata.get("journal", metadata.get("journal_name", "Unknown Journal")),
                metadata.get("doi", "None"),
                ", ".join(metadata.get("specialty", [])) if isinstance(metadata.get("specialty"), list) else metadata.get("system", "Other"),
                metadata.get("article_subtype", metadata.get("doc_type", metadata.get("type_of_article", "Other"))),
                "Yes", "No",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "",
                parsing_notes, "No"
            ]
            ws.append(row_data)
            wb.save(EXCEL_TRACKER_FILE)
            return
        except PermissionError:
            time.sleep(1.5)

# =====================================================================
# 🧠 FAIL-SAFE MODEL EXECUTION ENGINE
# =====================================================================
def execute_gemini_with_fallback(prompt, system_instruction, target_model, response_json=False):
    execution_matrix = [
        (PRIMARY_CLIENT, target_model),
        (BACKUP_CLIENT, target_model),
        (PRIMARY_CLIENT, MODEL_BACKUP)
    ]
    
    config = types.GenerateContentConfig(
        system_instruction=system_instruction,
        temperature=0.1,
        response_mime_type="application/json" if response_json else "text/plain"
    )
    
    last_error = None
    for client, model in execution_matrix:
        if not client:
            continue
        for attempt in range(3):
            try:
                response = client.models.generate_content(
                    model=model,
                    contents=prompt,
                    config=config
                )
                if response_json:
                    return json.loads(response.text.strip())
                return response.text.strip()
            except Exception as e:
                last_error = e
                time.sleep(3)
                
    raise RuntimeError(f"🚨 Critical Failure: API key layers exhausted. Error: {last_error}")

def generate_final_structured_payload(uploaded_file, category_tag):
    """Analyzes the file object natively via Google Files API into the new structured schema."""
    target_model = MODEL_GUIDELINES if category_tag == "guidelines" else MODEL_ARTICLES
    
    system_instruction = """You are an advanced clinical content extraction engine for hack.CCM's knowledge base. You will be given a medical PDF document (a research article or clinical guideline). Your job is to EXTRACT and STRUCTURE its content into the fixed schema below. Do NOT invent new clinical facts, do NOT omit facts present in the document, and do NOT re-interpret the content - only reorganize it.

CRITICAL RULES:
1. Output ONLY valid JSON. No preamble, no markdown fences, no commentary.
2. Determine doc_type: if the document is a guideline / clinical practice guideline, use the GUIDELINE schema. Otherwise (Review, RCT, Meta-analysis, Secondary Analysis, Observational study, Trial, Case Series, etc.) use the ARTICLE schema.
3. Collapse ALL nested heading levels (H3/H4/H5/H6) into the new schema's two-level-maximum structure: top-level sections with any deeper sub-headers folded into that section's "content"/"narrative" field as prose or inline bullets. Do not create more top-level sections than the original warrants - usually 4-10.
4. Extract "key_pearls" by scanning the document's key takeaways / clinical application content and converting its bullet points into 4-7 atomic, standalone pearls. If more than 7 bullets exist, select the 4-7 most clinically decisive ones (prioritize those with numbers, thresholds, or drug names).
5. For GUIDELINES specifically: scan for recommendation identifiers embedded in headers or text (e.g. "(R1.1.1)", "(Q3)") and use these to reconstruct "recommendation_blocks". Each major subsection becomes one recommendation_block, with its constituent statements broken into individual "recommendations" entries. If no clean granularity exists, create ONE recommendation entry per block with strength/evidence_grade set to null.
6. For GUIDELINES: if the document has a bedside protocol / step-by-step workflow section, map it directly into "bedside_protocol".
7. "specialty" must be an array of 1-3 controlled strings from: ["pulmonology","nephrology","hepatology","neurology","cardiology","infectious_disease","hematology","endocrinology","gastroenterology","toxicology","trauma","surgery","multi_system","pharmacology","rehabilitation"]. Split combined fields into multiple entries.
8. "one_line_summary" should be a single sentence max ~35 words extracted from the core summary.
9. "strengths_limitations" maps directly from the document's strengths/limitations content, condensed to 1-3 sentences.
10. Preserve "doi", "journal"/"issuing_bodies", "authors", and generate "id" as a slug from title + year if not already present.
11. Set "added_date" to today's date.
12. If genuinely unable to populate a field, use null or [] - never fabricate.
13. Do NOT use LaTeX formatting. Write out metrics, clearances, equations using plain text notation.
14. For ARTICLES: "evidence_level" must be exactly one of: "review", "rct", "meta_analysis", "secondary_analysis", "observational", "case_series", "narrative_review".
15. Enforce strict medical ground truth. Never generalize equations, target values, drug intervals, or data clearances.

ARTICLE SCHEMA:
{
  "id": "string - URL-safe slug from title + year",
  "doc_type": "article",
  "article_subtype": "review | rct | meta_analysis | secondary_analysis | observational | case_series | narrative_review",
  "title": "string - exact paper title",
  "authors": "string - first author + et al.",
  "journal": "string",
  "year": number,
  "doi": "string",
  "specialty": ["array of controlled strings, 1-3 items"],
  "tags": ["array of 3-8 free-text clinical keywords, lowercase"],
  "one_line_summary": "string, max ~35 words",
  "key_pearls": ["array of 4-7 atomic, standalone clinical pearls"],
  "evidence_level": "string, one of the controlled enum values",
  "sample_size": "number or null",
  "population": "string or null",
  "sections": [
    {
      "order": number,
      "heading": "string, plain title, no emoji, no markdown",
      "content": "string, plain prose with inline '- ' bullets for lists",
      "section_pearls": ["0-3 short pearls or empty array"]
    }
  ],
  "strengths_limitations": "string, bulletted points for strengths and limitations",
  "related_ids": [],
  "added_date": "YYYY-MM-DD"
}

For GUIDELINES:
{
  "id": "string - URL-safe slug from issuing body + topic + year",
  "doc_type": "guideline",
  "title": "string - exact guideline title",
  "issuing_bodies": ["array of society/organization acronyms, e.g. AHA, ACC, ESICM"],
  "year": number,
  "doi": "string",
  "specialty": ["array of controlled strings, 1-3 items"],
  "tags": ["array of 3-8 free-text clinical keywords, lowercase"],
  "one_line_summary": "string, max ~35 words",
  "key_pearls": ["array of 4-7 atomic, standalone clinical pearls"],
  "consensus_method": "string or null",
  "search_period": "string or null",
  "recommendation_blocks": [
    {
      "order": number,
      "topic": "string, plain title, no emoji, no markdown",
      "narrative": "string, preserve all numbers, p-values, ORs, trial names verbatim",
      "recommendations": [
        {
          "rec_id": "string or null",
          "statement": "string, clear directive",
          "strength": "strong | conditional | weak | expert_opinion | null",
          "evidence_grade": "string or null"
        }
      ]
    }
  ],
  "bedside_protocol": [
    {
      "step": number,
      "title": "string",
      "action": "string, 1-3 sentences"
    }
  ],
  "strengths_limitations": "string, bulletted points for strengths and limitations",
  "related_ids": [],
  "added_date": "YYYY-MM-DD"
}"""

    prompt_content = [
        uploaded_file,
        "Extract and structure the full clinical content of this document into the schema above. Preserve all numbers, thresholds, drug doses, trial endpoints, and evidence grades exactly as stated."
    ]

    print(f"  🧠 Analyzing payload architecture via {target_model}...")
    return execute_gemini_with_fallback(prompt_content, system_instruction, target_model, response_json=True)

# =====================================================================
# 📥 EXTRACTION PROCESSING LAYER
# =====================================================================
def process_single_pdf(file_path, category, processed_history):
    file_name = os.path.basename(file_path)
    if file_name in processed_history:
        return

    print(f"\n⚡ Ingesting target medical asset [{category.upper()}]: {file_name}")
    uploaded_file = None
    client_to_use = PRIMARY_CLIENT if PRIMARY_CLIENT else BACKUP_CLIENT
    
    try:
        print("  📤 Streaming document to Google Files API Hub...")
        uploaded_file = client_to_use.files.upload(file=file_path)
        
        while uploaded_file.state.name == "PROCESSING":
            time.sleep(2)
            uploaded_file = client_to_use.files.get(name=uploaded_file.name)
            
        if uploaded_file.state.name == "FAILED":
            raise ValueError("Google Files API dropped document serialization layers.")

        structured_payload = generate_final_structured_payload(uploaded_file, category)
        
        # Extract directory path components from new schema fields
        specialty_list = structured_payload.get("specialty", [])
        if isinstance(specialty_list, list) and len(specialty_list) > 0:
            clean_system = "".join(x for x in str(specialty_list[0]) if x.isalnum() or x in "._- ").strip()
        else:
            clean_system = "Other"
        
        doc_type = structured_payload.get("article_subtype", structured_payload.get("doc_type", "Other"))
        clean_type = "".join(x for x in str(doc_type) if x.isalnum() or x in "._- ").strip()
        
        sharded_output_dir = os.path.join(OUTPUT_DIR, clean_system, clean_type)
        os.makedirs(sharded_output_dir, exist_ok=True)
        
        base_name = os.path.splitext(file_name)[0]
        destination_json_path = os.path.join(sharded_output_dir, f"{base_name}.json")
        
        with open(destination_json_path, "w", encoding="utf-8") as jf:
            json.dump(structured_payload, jf, indent=2, ensure_ascii=False)
        print(f"  🌐 Ingestion matrix sharded to filesystem: {destination_json_path}")
        
        log_transaction_to_excel(file_name, structured_payload, "Success")
        processed_history.add(file_name)
        
        if os.path.exists(file_path):
            os.remove(file_path)

    except Exception as process_error:
        print(f"  ❌ Extraction Failure targeting ({file_name}): {process_error}")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            with open("./error_logs.txt", "a", encoding="utf-8") as ef:
                ef.write(f"[{timestamp}] File: {file_name} | Error: {str(process_error)}\n")
        except Exception:
            pass

    finally:
        if uploaded_file:
            try:
                client_to_use.files.delete(name=uploaded_file.name)
            except Exception:
                pass

# =====================================================================
# 📡 RUNTIME WATCHER LOOP POLLING
# =====================================================================
if __name__ == "__main__":
    initialize_system_paths()
    history_log = load_processed_files_from_excel()
    print("🚀 Advanced Medical Ingestion Engine Active...")
    print(f"  🩺 Specialties List: {len(ALLOWED_SPECIALTIES)} fields loaded.")
    print(f"  📑 Article Formats: {len(ALLOWED_TYPES)} variants loaded.")
    
    try:
        loop_counter = 0
        while True:
            loop_counter += 1
            history_log = load_processed_files_from_excel()
            
            if loop_counter % 12 == 1:
                print(f"⏱️ [Heartbeat] Monitoring watch folders... Time: {datetime.now().strftime('%H:%M:%S')}")
            
            for category, folder_path in SUB_DIRS.items():
                if os.path.exists(folder_path):
                    pdf_files = [f for f in os.listdir(folder_path) if f.endswith('.pdf')]
                    
                    for file_name in pdf_files:
                        path = os.path.join(folder_path, file_name)
                        if file_name in history_log:
                            continue
                            
                        try:
                            initial_size = os.path.getsize(path)
                            time.sleep(1.5)
                            if os.path.getsize(path) == initial_size:
                                process_single_pdf(path, category, history_log)
                        except Exception:
                            continue
            time.sleep(5)
    except KeyboardInterrupt:
        print("\n🛑 Ingestion loop cleanly terminated.")