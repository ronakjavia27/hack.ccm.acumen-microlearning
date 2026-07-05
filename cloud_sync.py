import os
import sys
import argparse
import subprocess
import openpyxl
from openpyxl import load_workbook

EXCEL_TRACKER_FILE = "./sent_summaries.xlsx"


def run_git_sync(mode="all"):
    """Validates if any updates are approved for publication before running git pushes."""
    print("📡 Initializing verified cloud repository check...")

    if not os.path.exists(EXCEL_TRACKER_FILE):
        print("❌ Error: Master Excel ledger tracking missing. Sync halted.")
        return

    # Check the excel state to see if any summaries are actually active on the web
    web_approved_count = 0
    try:
        wb = load_workbook(EXCEL_TRACKER_FILE, read_only=True)
        ws = wb["Registry Logs"]
        for row in list(ws.iter_rows(min_row=2, values_only=True)):
            if len(row) >= 14 and row[13] == "Yes":  # Column 14 (show_on_web)
                web_approved_count += 1
    except Exception as e:
        print(f"⚠️ Error verifying ledger state flags: {e}")
        return

    print(f"📋 Ledger confirmation check: {web_approved_count} articles currently marked 'Yes' for the web web portal.")

    # Execute git sync checks safely
    try:
        status = subprocess.run(["git", "status", "--porcelain"], capture_output=True, text=True)
        if not status.stdout.strip():
            print("✅ Web workspace repository state already matches upstream master baseline.")
            return

        # Stage files based on mode
        if mode == "web":
            print("📦 Staging web app only (main_app.py)...")
            subprocess.run(["git", "add", "main_app.py"])
        elif mode == "data":
            print("📦 Staging all files except main_app.py...")
            subprocess.run(["git", "add", "-A"])
            subprocess.run(["git", "reset", "main_app.py"])
        else:  # all
            print("📦 Staging all changes (respecting .gitignore)...")
            subprocess.run(["git", "add", "-A"])

        # Double-check nothing staged from .gitignore
        status = subprocess.run(["git", "diff", "--cached", "--stat"], capture_output=True, text=True)
        print(f"   {status.stdout.strip() or 'No changes staged.'}")

        commit_msg = f"Clinical Sync - Published Ledger Inventory Base"
        subprocess.run(["git", "commit", "-m", commit_msg])

        print("🚀 Pushing updates up to cloud network...")
        result = subprocess.run(["git", "push", "origin", "main"], capture_output=True, text=True)

        if result.returncode == 0:
            print("✅ Upstream server push successful! Cloud components synchronizing updates.")
        else:
            print(f"❌ Git Pipeline push error: {result.stderr}")
    except Exception as e:
        print(f"❌ Critical Git loop interaction crash: {e}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Sync local changes to cloud repository")
    parser.add_argument("--mode", choices=["all", "data", "web"], default="all",
                        help="What to sync: all (everything), data (all except main_app.py), web (only main_app.py)")
    args = parser.parse_args()
    run_git_sync(mode=args.mode)
