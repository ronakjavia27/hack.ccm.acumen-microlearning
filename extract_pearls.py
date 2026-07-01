import os
import sys
import re
import json
import argparse
from datetime import datetime
from google import genai
from google.genai import types

from dotenv import load_dotenv

script_dir = os.path.dirname(os.path.abspath(__file__))
env_path = os.path.join(script_dir, '.env')
load_dotenv(dotenv_path=env_path)

import openpyxl
from openpyxl import Workbook, load_workbook

# =====================================================================
# ⚙️ CONFIGURATIONS — change these as needed
# =====================================================================
MODE = "gemini"                          # "gemini" or "local"
MAX_PAPERS = 1                           # max JSONs to process per run; 0 = all
MODEL_PRIMARY = "gemini-3.5-flash"         # for pearl synthesis
MODEL_FALLBACK = "gemini-3.1-flash-lite"        # backup model


OUTPUT_DIR = "./output_files"
PEARLS_JSON = "./pearls.json"
PEARLS_TRACKER = "./pearls_processed.xlsx"

PRIMARY_API_KEY = os.getenv("PRIMARY_GEMINI_API_KEY")
BACKUP_API_KEY = os.getenv("BACKUP_GEMINI_API_KEY")

if not PRIMARY_API_KEY and not BACKUP_API_KEY:
    print("No API keys found. Local mode will be used as fallback.")
    MODE = "local"

PRIMARY_CLIENT = genai.Client(api_key=PRIMARY_API_KEY) if PRIMARY_API_KEY else None
BACKUP_CLIENT = genai.Client(api_key=BACKUP_API_KEY) if BACKUP_API_KEY else None

PEARLS_JSON_FIELDS = [
    "id", "timestamp", "source_paper", "doi",
    "author", "system", "type", "pearl", "remarks", "file_name", "topic"
]

def execute_gemini_with_fallback(prompt, system_instruction, target_model):
    execution_matrix = [
        (PRIMARY_CLIENT, target_model),
        (BACKUP_CLIENT, target_model),
        (PRIMARY_CLIENT, MODEL_FALLBACK)
    ]
    config = types.GenerateContentConfig(
        system_instruction=system_instruction,
        temperature=0.2,
        response_mime_type="application/json"
    )
    last_error = None
    for client, model in execution_matrix:
        if not client:
            continue
        for attempt in range(2):
            try:
                response = client.models.generate_content(
                    model=model,
                    contents=prompt,
                    config=config
                )
                raw = response.text.strip()
                raw = re.sub(r'^```(?:json)?\s*', '', raw)
                raw = re.sub(r'\s*```$', '', raw)
                return json.loads(raw)
            except Exception as e:
                last_error = e
    print(f"  Gemini call failed: {last_error}")
    return None

def load_tracker():
    if not os.path.exists(PEARLS_TRACKER):
        return set()
    try:
        wb = load_workbook(PEARLS_TRACKER, read_only=True)
        ws = wb["Tracker"]
        processed = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                processed.add(str(row[0]).strip())
        return processed
    except Exception:
        return set()

def update_tracker(file_name, pearl_count, mode):
    try:
        if os.path.exists(PEARLS_TRACKER):
            wb = load_workbook(PEARLS_TRACKER)
            ws = wb["Tracker"]
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Tracker"
            ws.append(["file_name", "timestamp_processed", "pearl_count", "mode"])
        ws.append([file_name, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), pearl_count, mode])
        wb.save(PEARLS_TRACKER)
    except Exception as e:
        print(f"  Warning: failed to update tracker: {e}")

def _atomic_write_json(file_path, data):
    tmp = file_path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    os.replace(tmp, file_path)


def load_existing_pearls():
    if not os.path.exists(PEARLS_JSON):
        return set(), 1
    try:
        with open(PEARLS_JSON, "r", encoding="utf-8") as f:
            rows = json.load(f)
        if not rows:
            return set(), 1
        ids = []
        source_papers = set()
        for r in rows:
            try:
                ids.append(int(r.get("id", 0)))
            except (ValueError, TypeError):
                pass
            sp = r.get("source_paper", "")
            if sp:
                source_papers.add(sp)
        next_id = max(ids) + 1 if ids else 1
        return source_papers, next_id
    except Exception:
        return set(), 1


def append_pearls_to_json(pearls, next_id):
    existing_rows = []
    if os.path.exists(PEARLS_JSON):
        try:
            with open(PEARLS_JSON, "r", encoding="utf-8") as f:
                existing_rows = json.load(f)
        except Exception:
            pass

    if existing_rows:
        all_ids = []
        for r in existing_rows:
            try:
                all_ids.append(int(r.get("id", 0)))
            except (ValueError, TypeError):
                pass
        next_id = (max(all_ids) + 1) if all_ids else 1

    now_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    all_rows = existing_rows + pearls

    for i, row in enumerate(all_rows):
        out = {}
        for key in PEARLS_JSON_FIELDS:
            val = row.get(key, "")
            if key == "id":
                val = str(i)
            elif key == "timestamp" and not val:
                val = now_ts
            out[key] = str(val) if val is not None else ""
        all_rows[i] = out

    _atomic_write_json(PEARLS_JSON, all_rows)
    return len(all_rows)

# =====================================================================
# 📥 LOCAL MODE — extract pearls from markdown using rules
# =====================================================================
def extract_pearls_local(markdown_text, metadata):
    pearls = []
    lines = markdown_text.split("\n")
    current_section = ""
    buffer = []

    for line in lines:
        stripped = line.strip()
        if stripped.startswith("### "):
            if buffer:
                text = " ".join(buffer).strip()
                if text and len(text) > 30:
                    pearls.append(text)
                buffer = []
            current_section = stripped
        elif stripped.startswith("-") or stripped.startswith("*"):
            content = stripped.lstrip("-* ").strip()
            if content and not content.startswith("**Detailed") and not content.startswith("**Granular"):
                buffer.append(content)
        elif stripped and not stripped.startswith("##") and not stripped.startswith("#"):
            if any(kw in stripped.lower() for kw in ["mg", "mmhg", "cmh2o", "%", "hr ", "p=", "p <", "p<", "ci ", "nnt", "nnh", "rr ", "or ", "hr="]):
                buffer.append(stripped)

    if buffer:
        text = " ".join(buffer).strip()
        if text and len(text) > 30:
            pearls.append(text)

    deduped = []
    seen = set()
    for p in pearls:
        key = p[:80].lower()
        if key not in seen:
            seen.add(key)
            deduped.append(p)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    result = []
    for p in deduped[:25]:
        result.append({
            "timestamp": timestamp,
            "source_paper": metadata.get("paper_name", ""),
            "doi": metadata.get("doi", ""),
            "author": metadata.get("primary_authors", ""),
            "system": metadata.get("system", ""),
            "type": metadata.get("type_of_article", ""),
            "pearl": p[:500],
            "remarks": "",
            "file_name": metadata.get("file_name", ""),
            "topic": metadata.get("topic", "")
        })
    return result

# =====================================================================
# 🤖 GEMINI MODE — use AI to synthesize pearls
# =====================================================================
def extract_pearls_gemini(markdown_text, metadata):
    prompt_text = f"""From the clinical summary below, extract all high-yield, evidence-based clinical pearls from the text. There is no minimum or maximum — extract exactly as many as the content warrants.

Each pearl must be:
- A single, specific, actionable takeaway
- Include exact numbers, doses, thresholds, or effect sizes where available
- Written in clear clinical language (1-3 sentences)

Return a JSON object with a "pearls" key containing an array of objects, each with "text" (the pearl string) and "topic" (1-3 topic keywords, comma-separated, e.g. "hemodynamics, vasopressors").
Example:
{{"pearls": [{{"text": "Start broad-spectrum antibiotics within 1 hour for septic shock (SSC 2026, strong recommendation, moderate quality evidence).", "topic": "sepsis, antibiotics"}}]}}

CLINICAL SUMMARY:
{markdown_text[:8000]}"""

    system_instruction = "You are an expert critical care clinician and medical educator tasked with extracting only high-yield, evidence-based clinical pearls from summarized medical articles and guidelines. From the provided text, output a numbered list of concise pearls that: directly impact clinical decision-making, bedside management, or exam-level reasoning (not generic advice or truisms); are specific and concrete (thresholds, cutoffs, dosing ranges, timing, risk modifiers, diagnostic criteria, prognostic markers, or management algorithms); are traceable to evidence (RCTs, meta-analyses, strong guideline recommendations, or consistently replicated observational data) when such information is present in the text; avoid generic statements such as “always assess level of evidence” or “more research is needed”; and preferentially capture practice-changing points, nuances (subgroups, exceptions, contraindications), and clear if–then conditions. Format: output only a numbered list, each pearl 1–2 sentences, maximally information-dense and self-contained; if no qualifying pearls exist, respond exactly: “No high-yield, evidence-based clinical pearls identifiable from this excerpt. Also tag each pearl with 1-3 topic keywords (e.g., 'hemodynamics, vasopressors', 'ventilation, ARDS', 'antibiotics, sepsis') as a 'topic' field."

    result = execute_gemini_with_fallback(prompt_text, system_instruction, MODEL_PRIMARY)
    if result and "pearls" in result and isinstance(result["pearls"], list):
        raw_pearls = result["pearls"]
    else:
        print("  Gemini returned unexpected format, falling back to local extraction.")
        return extract_pearls_local(markdown_text, metadata)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    output = []
    for p in raw_pearls[:25]:
        if isinstance(p, dict):
            text = p.get("text", "").strip()
            topic = p.get("topic", "").strip()
            if len(text) > 15:
                output.append({
                    "timestamp": timestamp,
                    "source_paper": metadata.get("paper_name", ""),
                    "doi": metadata.get("doi", ""),
                    "author": metadata.get("primary_authors", ""),
                    "system": metadata.get("system", ""),
                    "type": metadata.get("type_of_article", ""),
                    "pearl": text[:500],
                    "remarks": "",
                    "file_name": metadata.get("file_name", ""),
                    "topic": topic
                })
        elif isinstance(p, str) and len(p) > 15:
            output.append({
                "timestamp": timestamp,
                "source_paper": metadata.get("paper_name", ""),
                "doi": metadata.get("doi", ""),
                "author": metadata.get("primary_authors", ""),
                "system": metadata.get("system", ""),
                "type": metadata.get("type_of_article", ""),
                "pearl": p.strip()[:500],
                "remarks": "",
                "file_name": metadata.get("file_name", ""),
                "topic": ""
            })
    return output

# =====================================================================
# 🚀 MAIN
# =====================================================================
def main():
    parser = argparse.ArgumentParser(description="Extract clinical pearls from JSON summaries.")
    parser.add_argument("--mode", choices=["gemini", "local"], default=MODE,
                        help=f"Extraction mode (default: {MODE})")
    parser.add_argument("--max", type=int, default=None,
                        help=f"Max papers to process (default: {MAX_PAPERS}; 0 = all)")
    args = parser.parse_args()

    mode = args.mode
    max_papers = args.max if args.max is not None else MAX_PAPERS

    print(f"🧠 hack.CCM Clinical Pearl Extractor")
    print(f"  Mode: {mode}")
    print(f"  Max papers: {'all' if max_papers == 0 else max_papers}")
    print(f"  Model: {MODEL_PRIMARY if mode == 'gemini' else 'N/A (local)'}")
    print()

    processed_files = load_tracker()
    existing_papers, next_id = load_existing_pearls()

    json_files = []
    for root, dirs, files in os.walk(OUTPUT_DIR):
        for fname in files:
            if fname.endswith(".json"):
                json_files.append(os.path.join(root, fname))

    print(f"📁 Found {len(json_files)} total JSON files")
    print(f"  Already processed: {len(processed_files)} files")
    print(f"  Already in CSV: {len(existing_papers)} papers")

    to_process = []
    for fpath in json_files:
        fname = os.path.basename(fpath)
        if fname in processed_files:
            continue
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                payload = json.load(f)
            paper_name = payload.get("paper_name", "")
            if paper_name in existing_papers:
                continue
            to_process.append((fpath, payload))
        except Exception:
            continue

    if max_papers > 0:
        to_process = to_process[:max_papers]

    print(f"  To process this run: {len(to_process)} files")
    print()

    if not to_process:
        print("✅ No new files to process. Exiting.")
        return

    total_pearls = 0
    for idx, (fpath, payload) in enumerate(to_process, start=1):
        fname = os.path.basename(fpath)
        payload["file_name"] = fname
        paper_name = payload.get("paper_name", fname)
        print(f"[{idx}/{len(to_process)}] {paper_name}")

        markdown = payload.get("clinical_summary_markdown", "")
        if not markdown or len(markdown) < 50:
            print(f"  ⏭️  Skipping (no content)")
            update_tracker(fname, 0, mode)
            continue

        if mode == "gemini":
            pearls = extract_pearls_gemini(markdown, payload)
        else:
            pearls = extract_pearls_local(markdown, payload)

        if not pearls:
            print(f"  ⏭️  No pearls extracted")
            update_tracker(fname, 0, mode)
            continue

        next_id = append_pearls_to_json(pearls, next_id)
        update_tracker(fname, len(pearls), mode)
        total_pearls += len(pearls)
        print(f"  ✅ {len(pearls)} pearls saved")

    print()
    print(f"🎉 Done! {total_pearls} new pearls written to {PEARLS_JSON}")

    # Validate the JSON is consistent
    try:
        with open(PEARLS_JSON, "r", encoding="utf-8") as f:
            rows = json.load(f)
        if rows and not all(k in rows[0] for k in PEARLS_JSON_FIELDS):
            missing = [k for k in PEARLS_JSON_FIELDS if k not in rows[0]]
            print(f"  ⚠️  Missing fields: {missing}")
        else:
            print(f"  ✅ JSON validated: {len(rows)} entries")
    except Exception as e:
        print(f"  ⚠️  JSON validation failed: {e}")

if __name__ == "__main__":
    main()
