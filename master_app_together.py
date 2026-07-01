#!/usr/bin/env python3
"""
master_app_together.py - hack.CCM medical PDF ingestion engine
using Together AI (DeepSeek V4 Pro) for structured JSON extraction.

Usage:
    pip install together openai pypdf openpyxl python-dotenv
    python master_app_together.py

Watches input_pdfs/{articles,guidelines,other}/ for new PDFs,
extracts text via PyPDF2, sends to DeepSeek V4 Pro on Together AI,
saves structured JSON to output_files/ and logs to sent_summaries.xlsx.
"""

import os
import sys
import time
import json
import re
import argparse
from copy import deepcopy
from datetime import datetime
from pypdf import PdfReader
from together import Together
from openai import OpenAI

from dotenv import load_dotenv

script_dir = os.path.dirname(os.path.abspath(__file__))
env_path = os.path.join(script_dir, '.env')
load_dotenv(dotenv_path=env_path)

import openpyxl
from openpyxl import Workbook, load_workbook

# =====================================================================
# ⚙️ CONFIGURATION
# =====================================================================
BASE_INPUT_DIR = "./input_pdfs"
SUB_DIRS = {
    "articles": os.path.join(BASE_INPUT_DIR, "articles"),
    "guidelines": os.path.join(BASE_INPUT_DIR, "guidelines"),
    "other": os.path.join(BASE_INPUT_DIR, "other")
}

OUTPUT_DIR = "./output_files"
EXCEL_TRACKER_FILE = "./sent_summaries.xlsx"
SPECIALTIES_FILE = "./specialties.txt"
ARTICLE_TYPES_FILE = "./article_types.txt"

# Model IDs
MODEL_TOGETHER_PRO = "deepseek-ai/DeepSeek-V4-Pro"
MODEL_TOGETHER_FLASH = "deepseek-ai/DeepSeek-V4-Pro"
MODEL_DEEPSEEK_DIRECT = "deepseek-v4-pro"

# API Keys
TOGETHER_API_KEY = os.getenv("TOGETHER_API_KEY")
DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY")

if not TOGETHER_API_KEY and not DEEPSEEK_API_KEY:
    print("CRITICAL: No API keys found. Set TOGETHER_API_KEY or DEEPSEEK_API_KEY in .env")
    sys.exit(1)

TOGETHER_CLIENT = Together(api_key=TOGETHER_API_KEY, timeout=300) if TOGETHER_API_KEY else None
DEEPSEEK_CLIENT = OpenAI(api_key=DEEPSEEK_API_KEY, base_url="https://api.deepseek.com") if DEEPSEEK_API_KEY else None

MAX_RETRIES = 3
RETRY_DELAY = 10  # seconds between retries; DeepSeek V4 Pro can take 60-120s per call
TEMPERATURE = 0.3
MAX_TOKENS = 16384
CHUNK_SIZE = 400000  # chars per chunk (~100K tokens); docs exceeding this get split
CHUNK_OVERLAP = 3000  # chars of overlap between chunks to avoid content loss at seams

# =====================================================================
# 📋 VOCABULARY MANAGEMENT
# =====================================================================
def load_allowed_vocabulary(file_path, default_list):
    if not os.path.exists(file_path):
        with open(file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(default_list))
        return default_list
    with open(file_path, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]

DEFAULT_SPECIALTIES = ["Critical Care Medicine", "Cardiovascular", "Neurology", "Nephrology", "Pulmonology", "Other"]
DEFAULT_TYPES = ["Guideline", "Review", "Meta-Analysis", "Trial", "Other"]

ALLOWED_SPECIALTIES = load_allowed_vocabulary(SPECIALTIES_FILE, DEFAULT_SPECIALTIES)
ALLOWED_TYPES = load_allowed_vocabulary(ARTICLE_TYPES_FILE, DEFAULT_TYPES)


def build_specialty_map(allowed):
    m = {s.lower(): s for s in allowed}
    m.update({
        "infectious_disease": m.get("infectious diseases", "Infectious Diseases"),
        "multi_system": m.get("multisystem", "Multisystem"),
        "multisystem": m.get("multisystem", "Multisystem"),
        "obstetrics_and_gynecology": m.get("obstetrics and gynecology", "Obstetrics and Gynecology"),
        "cardio": m.get("cardiology", "Cardiology"),
        "neuro": m.get("neurology", "Neurology"),
        "nephro": m.get("nephrology", "Nephrology"),
        "pulmo": m.get("pulmonology", "Pulmonology"),
        "gi": m.get("gastroenterology", "Gastroenterology"),
        "heme": m.get("hematology", "Hematology"),
        "onc": m.get("oncology", "Oncology"),
    })
    return m


def normalize_specialty(specialty_list, spec_map):
    if not isinstance(specialty_list, list) or not specialty_list:
        return "Other"
    raw = str(specialty_list[0]).strip().lower().replace("_", " ").replace("-", " ")
    mapped = spec_map.get(raw, "Other")
    return "".join(x for x in str(mapped) if x.isalnum() or x in "._- ").strip()


def normalize_type(payload, allowed_types):
    low_types = {}
    for t in allowed_types:
        low_types[t.lower().replace("_", "-").replace(" ", "-")] = t
    doc_type = (payload.get("doc_type") or "").lower()
    if doc_type == "guideline":
        payload["article_subtype"] = "Guideline"
        return "Guideline"
    subtype = (payload.get("article_subtype") or "").strip()
    normalized = subtype.lower().replace("_", "-").replace(" ", "-")
    if normalized in low_types:
        payload["article_subtype"] = low_types[normalized]
        return low_types[normalized]
    payload["article_subtype"] = "Other"
    return "Other"


def apply_markdown_emphasis(text):
    if not text:
        return text
    parts = re.split(r'(\*\*[^*]+\*\*)', text)
    for i, part in enumerate(parts):
        if i % 2 == 1:
            continue
        parts[i] = re.sub(
            r'(\d+(?:\.\d+)?)\s*(mg|mcg|g|mL|L|mmHg|cmH2O|%|mmol|mEq|IU|U|kg|hr|min|mg/kg|mcg/kg|mEq/L|mmol/L|mg/dL|IU/kg/hr)\b',
            r'**\1 \2**',
            parts[i], flags=re.IGNORECASE
        )
        parts[i] = re.sub(
            r'\b(significant(?:ly)?|recommended|contraindicated|critical(?:ly)?|essential|pivotal|superior|inferior|equivalent|mandatory|absolute|mortality|survival|prognosis|notable)\b',
            lambda m: f'**{m.group(0)}**',
            parts[i], flags=re.IGNORECASE
        )
    return ''.join(parts)


def enrich_payload_with_markdown(payload):
    for key in ['one_line_summary', 'strengths_limitations']:
        if isinstance(payload.get(key), str):
            payload[key] = apply_markdown_emphasis(payload[key])
    for s in payload.get('sections', []):
        if isinstance(s.get('content'), str):
            s['content'] = apply_markdown_emphasis(s['content'])
    for b in payload.get('recommendation_blocks', []):
        if isinstance(b.get('narrative'), str):
            b['narrative'] = apply_markdown_emphasis(b['narrative'])
        for r in b.get('recommendations', []):
            if isinstance(r.get('statement'), str):
                r['statement'] = apply_markdown_emphasis(r['statement'])
    for i, p in enumerate(payload.get('key_pearls', [])):
        if isinstance(p, str):
            payload['key_pearls'][i] = apply_markdown_emphasis(p)
    return payload


# =====================================================================
# 📊 EXCEL TRACKING
# =====================================================================
EXCEL_HEADERS = [
    "Serial Number", "File Name", "Paper/Guideline Name", "Primary Authors",
    "Journal Name", "DOI", "Year", "System", "Type of Article", "MD Generated",
    "Email Pushed", "Summary Saved Date", "Email Pushed Date", "Parsing Notes", "show_on_web"
]

def initialize_system_paths():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    for folder_path in SUB_DIRS.values():
        os.makedirs(folder_path, exist_ok=True)

    if not os.path.exists(EXCEL_TRACKER_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registry Logs"
        ws.append(EXCEL_HEADERS)
        wb.save(EXCEL_TRACKER_FILE)
        print(f"Excel tracker initialized at {EXCEL_TRACKER_FILE}")
    else:
        migrate_ledger_schema()


def migrate_ledger_schema():
    """Add Year column to existing ledger if missing."""
    try:
        wb = load_workbook(EXCEL_TRACKER_FILE)
        ws = wb["Registry Logs"]
        headers = [cell.value for cell in ws[1]]
        if "Year" not in headers:
            col_idx = len(headers) + 1
            ws.cell(row=1, column=col_idx, value="Year")
            wb.save(EXCEL_TRACKER_FILE)
            print(f"  Migrated ledger: added Year column")
    except Exception:
        pass

def load_processed_files_from_excel():
    if not os.path.exists(EXCEL_TRACKER_FILE):
        return set()
    try:
        wb = load_workbook(EXCEL_TRACKER_FILE, read_only=True)
        ws = wb["Registry Logs"]
        processed = set()
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if len(row) > 1 and row[1]:
                processed.add(str(row[1]).strip())
        return processed
    except Exception:
        return set()

def log_transaction_to_excel(file_name, metadata, parsing_notes="Success"):
    retries = 5
    for attempt in range(retries):
        try:
            wb = load_workbook(EXCEL_TRACKER_FILE)
            ws = wb["Registry Logs"]
            next_serial = ws.max_row

            row_data = [
                next_serial, file_name,
                metadata.get("title", metadata.get("paper_name", "Unknown Title")),
                metadata.get("authors", metadata.get("primary_authors", "Unknown Authors")),
                metadata.get("journal", metadata.get("journal_name", "Unknown Journal")),
                metadata.get("doi", "None"),
                metadata.get("year", ""),
                ", ".join(metadata.get("specialty", [])) if isinstance(metadata.get("specialty"), list) else metadata.get("system", "Other"),
                metadata.get("article_subtype", metadata.get("doc_type", metadata.get("type_of_article", "Other"))),
                "Yes", "No",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "",
                parsing_notes, "No"
            ]
            ws.append(row_data)
            wb.save(EXCEL_TRACKER_FILE)
            return
        except PermissionError:
            time.sleep(1.5)

# =====================================================================
# 🧠 STRUCTURED SYSTEM PROMPT (from migrate_jsons.py)
# =====================================================================
SYSTEM_PROMPT = """You are an advanced clinical content extraction engine for hack.CCM's knowledge base. You will be given the extracted text of a medical PDF document (a research article or clinical guideline). Your job is to EXTRACT and STRUCTURE its content into the fixed schema below. Do NOT invent new clinical facts, do NOT omit facts present in the document, and do NOT re-interpret the content - only reorganize it.

CRITICAL RULES:
1. Output ONLY valid JSON. No preamble, no markdown fences, no commentary.
2. Determine doc_type: if the document is a guideline / clinical practice guideline, use the GUIDELINE schema. Otherwise (Review, RCT, Meta-analysis, Secondary Analysis, Observational study, Trial, Case Series, etc.) use the ARTICLE schema.
3. Collapse ALL nested heading levels (H3/H4/H5/H6) into the new schema's two-level-maximum structure: top-level sections with any deeper sub-headers folded into that section's "content"/"narrative" field as structured bullet points (-), no prose paragraphs. Do not create more top-level sections than the original warrants - usually 4-10.
4. Extract "key_pearls" by scanning the document's key takeaways / clinical application content and converting its bullet points into 4-7 atomic, standalone pearls. If more than 7 bullets exist, select the 4-7 most clinically decisive ones (prioritize those with numbers, thresholds, or drug names).
5. For GUIDELINES specifically: scan for recommendation identifiers embedded in headers or text (e.g. "(R1.1.1)", "(Q3)") and use these to reconstruct "recommendation_blocks". Each major subsection becomes one recommendation_block, with its constituent statements broken into individual "recommendations" entries. If no clean granularity exists, create ONE recommendation entry per block with strength/evidence_grade set to null.
6. For GUIDELINES: if the document has a bedside protocol / step-by-step workflow section, map it directly into "bedside_protocol".
7. "specialty" must be an array of 1-3 controlled strings from: ["pulmonology","nephrology","hepatology","neurology","cardiology","infectious_disease","hematology","endocrinology","gastroenterology","toxicology","trauma","surgery","multi_system","pharmacology","rehabilitation"]. Split combined fields into multiple entries.
8. "one_line_summary" should be a single sentence max ~35 words extracted from the core summary.
9. "strengths_limitations" maps directly from the document's strengths/limitations content, condensed to 1-3 sentences.
10. Preserve "doi", "journal"/"issuing_bodies", "authors", and generate "id" as a slug from title + year if not already present.
11. Set "added_date" to today's date.
12. If genuinely unable to populate a field, use null or [] - never fabricate.
13. Do NOT use LaTeX formatting. Write out metrics, clearances, equations using plain text notation.
14. For ARTICLES: "evidence_level" must be exactly one of: "review", "rct", "meta_analysis", "secondary_analysis", "observational", "case_series", "narrative_review".
15. Enforce strict medical ground truth. Never generalize equations, target values, drug intervals, or data clearances.
16. Format ALL content fields as structured bullet points (- item per finding), never as prose paragraphs. Each bullet captures one atomic finding, data point, or recommendation. Use markdown formatting for clinical emphasis: wrap key numbers, thresholds, drug names, lab values, and clinical endpoints in **bold**; use *italic* for technical terms and statistical measures. Apply this within sections[].content, narrative, recommendation statements, and key_pearls.
17. Extract all drugs mentioned with their doses, routes, frequencies, indications, and key adverse effects into the "drugs_doses" array. One entry per drug-context pair. If a drug appears in multiple indications, include separate entries. If none, use [].

ARTICLE SCHEMA:
{
  "id": "string - URL-safe slug from title + year",
  "doc_type": "article",
  "article_subtype": "review | rct | meta_analysis | secondary_analysis | observational | case_series | narrative_review",
  "title": "string - exact paper title",
  "authors": "string - first author + et al.",
  "journal": "string",
  "year": number,
  "doi": "string",
  "specialty": ["array of controlled strings, 1-3 items"],
  "tags": ["array of 3-8 free-text clinical keywords, lowercase"],
  "one_line_summary": "string, max ~35 words",
  "key_pearls": ["array of 4-7 atomic, standalone clinical pearls"],
  "evidence_level": "string, one of the controlled enum values",
  "sample_size": "number or null",
  "population": "string or null",
  "sections": [
    {
      "order": number,
      "heading": "string, plain title, no emoji, no markdown",
      "content": "string, bullet points (- item per key finding), no prose paragraphs",
      "section_pearls": ["0-3 short pearls or empty array"]
    }
  ],
  "strengths_limitations": "string, bulletted points for strengths and limitations",
  "drugs_doses": [
    {
      "drug": "string - generic drug name",
      "dose": "string - dose, route, frequency, duration",
      "indication": "string - clinical context",
      "adverse_effects": "string - key side effects, contraindications, monitoring"
    }
  ],
  "related_ids": [],
  "added_date": "YYYY-MM-DD"
}

For GUIDELINES:
{
  "id": "string - URL-safe slug from issuing body + topic + year",
  "doc_type": "guideline",
  "title": "string - exact guideline title",
  "issuing_bodies": ["array of society/organization acronyms, e.g. AHA, ACC, ESICM"],
  "year": number,
  "doi": "string",
  "specialty": ["array of controlled strings, 1-3 items"],
  "tags": ["array of 3-8 free-text clinical keywords, lowercase"],
  "one_line_summary": "string, max ~35 words",
  "key_pearls": ["array of 4-7 atomic, standalone clinical pearls"],
  "consensus_method": "string or null",
  "search_period": "string or null",
  "recommendation_blocks": [
    {
      "order": number,
      "topic": "string, plain title, no emoji, no markdown",
      "narrative": "string, bullet points; preserve all numbers, p-values, ORs, trial names verbatim",
      "recommendations": [
        {
          "rec_id": "string or null",
          "statement": "string, clear directive",
          "strength": "strong | conditional | weak | expert_opinion | null",
          "evidence_grade": "string or null"
        }
      ]
    }
  ],
  "bedside_protocol": [
    {
      "step": number,
      "title": "string",
      "action": "string, 1-3 sentences"
    }
  ],
  "strengths_limitations": "string, bulletted points for strengths and limitations",
  "drugs_doses": [
    {
      "drug": "string - generic drug name",
      "dose": "string - dose, route, frequency, duration",
      "indication": "string - clinical context",
      "adverse_effects": "string - key side effects, contraindications, monitoring"
    }
  ],
  "related_ids": [],
  "added_date": "YYYY-MM-DD"
}"""

# =====================================================================
# 📡 TOGETHER AI / DEEPSEEK API EXECUTION
# =====================================================================
def call_together(client, model, system_prompt, document_text):
    """Single API call to Together AI or Direct DeepSeek."""
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": f"Extract and structure the following clinical document into JSON format. Return ONLY valid JSON following the schema from the system prompt:\n\n{document_text}"}
    ]

    kwargs = {
        "model": model,
        "messages": messages,
        "response_format": {"type": "json_object"},
        "temperature": TEMPERATURE,
        "max_tokens": MAX_TOKENS,
    }
    # Only Together AI supports reasoning control; skip for direct DeepSeek API
    if hasattr(client, '_client_config') or 'together' in str(type(client)).lower():
        kwargs["reasoning"] = {"enabled": False}

    response = client.chat.completions.create(**kwargs)

    raw = response.choices[0].message.content
    if raw is None or raw.strip() == "":
        raise ValueError("Empty response from model")

    raw = raw.strip()
    # Fallback: strip markdown fences if they appear despite JSON mode
    raw = re.sub(r'^```(?:json)?\s*\n?', '', raw)
    raw = re.sub(r'\n?\s*```$', '', raw)

    return json.loads(raw)


def execute_with_fallback(document_text, category_tag):
    """Try Together AI Pro -> Together AI Flash -> Direct DeepSeek API."""
    models_tog = [MODEL_TOGETHER_PRO, MODEL_TOGETHER_FLASH]

    last_error = None

    def is_retryable(e):
        msg = str(e).lower()
        if 'timeout' in msg or 'timed out' in msg:
            return True
        if 'rate limit' in msg or '429' in msg or 'resource_exhausted' in msg:
            return True
        if '503' in msg or '502' in msg or 'service unavailable' in msg:
            return True
        if 'internal server' in msg or '500' in msg:
            return True
        if 'empty response' in msg:
            return True
        return False

    # Phase 1: Together AI models
    if TOGETHER_CLIENT:
        for model in models_tog:
            for attempt in range(MAX_RETRIES):
                try:
                    print(f"    Together AI: {model} (attempt {attempt + 1}/{MAX_RETRIES})")
                    return call_together(TOGETHER_CLIENT, model, SYSTEM_PROMPT, document_text)
                except json.JSONDecodeError as e:
                    last_error = e
                    print(f"    [X] JSON parse error: {e}")
                    break
                except Exception as e:
                    last_error = e
                    if is_retryable(e) and attempt < MAX_RETRIES - 1:
                        wait = RETRY_DELAY * (attempt + 1)
                        print(f"    [!] {e}")
                        print(f"    Retrying in {wait}s...")
                        time.sleep(wait)
                    else:
                        reason = "non-retryable" if not is_retryable(e) else "retries exhausted"
                        print(f"    [X] {model} failed ({reason}): {e}")
                        break

    # Phase 2: Direct DeepSeek API fallback
    if DEEPSEEK_CLIENT:
        try:
            print(f"    Direct DeepSeek: {MODEL_DEEPSEEK_DIRECT}")
            return call_together(DEEPSEEK_CLIENT, MODEL_DEEPSEEK_DIRECT, SYSTEM_PROMPT, document_text)
        except Exception as e:
            last_error = e
            print(f"    [X] Direct DeepSeek failed: {e}")

    raise last_error or RuntimeError("All models exhausted")


# =====================================================================
# 🔪 TEXT CHUNKING & PROGRAMMATIC MERGE
# =====================================================================
def chunk_text(text, chunk_size=CHUNK_SIZE, overlap=CHUNK_OVERLAP):
    """Split text at paragraph boundaries. Each chunk <= chunk_size with overlap."""
    if len(text) <= chunk_size:
        return [text]

    result = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        if end >= len(text):
            result.append(text[start:])
            break
        # Find nearest paragraph break before end
        boundary = text.rfind('\n\n', start, end)
        if boundary == -1 or (boundary - start) < chunk_size // 4:
            boundary = text.rfind('\n', start, end)
            if boundary == -1 or (boundary - start) < chunk_size // 4:
                boundary = text.rfind(' ', start, end)
                if boundary == -1:
                    boundary = end
        chunk = text[start:boundary]
        result.append(chunk)
        # Next chunk starts before the boundary to include overlap
        start = max(boundary - overlap, boundary - chunk_size // 4)
    return result


def merge_chunks_programmatically(results):
    """Merge chunk outputs without an extra API call.
    Keeps metadata from first chunk, concatenates list-type fields.
    """
    if not results:
        return None
    if len(results) == 1:
        return deepcopy(results[0])

    merged = deepcopy(results[0])

    for chunk_result in results[1:]:
        # Merge sections
        existing_sections = merged.get("sections", [])
        offset = len(existing_sections)
        for s in chunk_result.get("sections", []):
            s["order"] = s.get("order", 0) + offset
            existing_sections.append(s)

        # Merge recommendation_blocks
        existing_blocks = merged.get("recommendation_blocks", [])
        offset = len(existing_blocks)
        for b in chunk_result.get("recommendation_blocks", []):
            b["order"] = b.get("order", 0) + offset
            existing_blocks.append(b)

        # Merge bedside_protocol
        existing_steps = merged.get("bedside_protocol", [])
        offset = len(existing_steps)
        for s in chunk_result.get("bedside_protocol", []):
            s["step"] = s.get("step", 0) + offset
            existing_steps.append(s)

        # Merge key_pearls (dedup by lower case)
        existing_pearls = merged.setdefault("key_pearls", [])
        seen = set(p.strip().lower() for p in existing_pearls)
        for p in chunk_result.get("key_pearls", []):
            key = p.strip().lower()
            if key and key not in seen:
                existing_pearls.append(p)
                seen.add(key)

        # Merge tags (unique union)
        existing_tags = merged.setdefault("tags", [])
        seen_tags = set(t.strip().lower() for t in existing_tags)
        for t in chunk_result.get("tags", []):
            key = t.strip().lower()
            if key and key not in seen_tags:
                existing_tags.append(t)
                seen_tags.add(key)

        # Concatenate strengths_limitations
        sl_existing = merged.get("strengths_limitations", "") or ""
        sl_chunk = chunk_result.get("strengths_limitations", "") or ""
        if sl_chunk and sl_chunk not in sl_existing:
            merged["strengths_limitations"] = (sl_existing + "\n" + sl_chunk).strip()

    return merged


# =====================================================================
# 📥 SINGLE PDF PROCESSING (with chunking)
# =====================================================================
def process_single_pdf(file_path, category, processed_history):
    file_name = os.path.basename(file_path)
    if file_name in processed_history:
        return

    print(f"\nIngesting [{category.upper()}]: {file_name}")

    try:
        # Extract text from PDF
        print("  Reading PDF...")
        reader = PdfReader(file_path)
        chunks = [page.extract_text() or "" for page in reader.pages]
        full_text = "\n".join(chunks).strip()

        if len(full_text) < 150:
            raise ValueError("Extracted text too short or unreadable")

        # Chunk if document exceeds per-chunk limit
        text_chunks = chunk_text(full_text, CHUNK_SIZE, CHUNK_OVERLAP)

        if len(text_chunks) == 1:
            print(f"  Extracted {len(full_text)} chars, calling DeepSeek V4 Pro...")
            structured_payload = execute_with_fallback(full_text, category)
        else:
            print(f"  Document large ({len(full_text)} chars). Splitting into {len(text_chunks)} chunks...")
            chunk_results = []
            for i, chunk in enumerate(text_chunks):
                print(f"  Chunk {i + 1}/{len(text_chunks)} ({len(chunk)} chars)...")
                chunk_result = execute_with_fallback(chunk, category)
                chunk_results.append(chunk_result)
            structured_payload = merge_chunks_programmatically(chunk_results)
            print(f"  Merged {len(chunk_results)} chunks into single JSON")

        # Normalize specialty against controlled vocabulary
        spec_map = build_specialty_map(ALLOWED_SPECIALTIES)
        clean_system = normalize_specialty(structured_payload.get("specialty", []), spec_map)

        # Normalize article type (invalid → "Other")
        clean_type_raw = normalize_type(structured_payload, ALLOWED_TYPES)
        clean_type = "".join(x for x in str(clean_type_raw) if x.isalnum() or x in "._- ").strip() or "Other"

        # Enrich text fields with markdown emphasis (bold/italics)
        structured_payload = enrich_payload_with_markdown(structured_payload)

        sharded_output_dir = os.path.join(OUTPUT_DIR, clean_system, clean_type)
        os.makedirs(sharded_output_dir, exist_ok=True)

        base_name = os.path.splitext(file_name)[0]
        destination_json_path = os.path.join(sharded_output_dir, f"{base_name}.json")

        with open(destination_json_path, "w", encoding="utf-8") as jf:
            json.dump(structured_payload, jf, indent=2, ensure_ascii=False)
        print(f"  Saved: {destination_json_path}")

        # Log with specialty clamped to match directory path
        log_meta = dict(structured_payload)
        log_meta["specialty"] = [clean_system]
        log_transaction_to_excel(file_name, log_meta, "Success")
        processed_history.add(file_name)

        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"  Deleted source PDF: {file_name}")

    except Exception as process_error:
        print(f"  [X] Failed: {process_error}")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            with open("./error_logs.txt", "a", encoding="utf-8") as ef:
                ef.write(f"[{timestamp}] File: {file_name} | Error: {str(process_error)}\n")
        except Exception:
            pass


# =====================================================================
# 📡 DIRECTORY WATCHER LOOP
# =====================================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="hack.CCM - Together AI Ingestion Engine")
    parser.add_argument("--max", type=int, default=0, help="Max files to process (0 = all)")
    args = parser.parse_args()
    max_files = args.max

    initialize_system_paths()
    history_log = load_processed_files_from_excel()
    print("hack.CCM - Together AI Ingestion Engine")
    print(f"  Models: {MODEL_TOGETHER_PRO} / {MODEL_TOGETHER_FLASH}")
    print(f"  Fallback: Direct DeepSeek {MODEL_DEEPSEEK_DIRECT}")
    print(f"  Chunk size: {CHUNK_SIZE:,} chars")
    if max_files > 0:
        print(f"  Max files: {max_files}")
    print(f"  Watching: {BASE_INPUT_DIR}/{{articles,guidelines,other}}/")
    print()

    try:
        loop_counter = 0
        files_processed = 0
        while True:
            loop_counter += 1
            history_log = load_processed_files_from_excel()

            if loop_counter % 12 == 1:
                print(f"Heartbeat - {datetime.now().strftime('%H:%M:%S')}")

            for category, folder_path in SUB_DIRS.items():
                if max_files > 0 and files_processed >= max_files:
                    break
                if os.path.exists(folder_path):
                    pdf_files = [f for f in os.listdir(folder_path) if f.endswith('.pdf')]

                    for file_name in pdf_files:
                        if max_files > 0 and files_processed >= max_files:
                            break
                        path = os.path.join(folder_path, file_name)
                        if file_name in history_log:
                            continue

                        try:
                            initial_size = os.path.getsize(path)
                            time.sleep(1.5)
                            if os.path.getsize(path) == initial_size:
                                process_single_pdf(path, category, history_log)
                                files_processed += 1
                        except Exception:
                            continue
                if max_files > 0 and files_processed >= max_files:
                    break
            if max_files > 0 and files_processed >= max_files:
                print(f"\nReached --max limit ({max_files}). Exiting.")
                break
            time.sleep(5)
    except KeyboardInterrupt:
        print("\nStopped.")
