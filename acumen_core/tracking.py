"""
tracking.py - Atomic JSON/Excel read/write utilities for trackers.
"""

import os
import json
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook

from acumen_core.config import (
    EXCEL_TRACKER_FILE,
    JSON_TRACKER_FILE,
    REMOVED_TRACKER_FILE,
    EXCEL_HEADERS,
)


def _atomic_write_json(file_path, data):
    """Write JSON atomically using tmp file + replace."""
    tmp = file_path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    os.replace(tmp, file_path)


def load_json_safe(file_path, default=None):
    """Load JSON file, returning default on error/missing."""
    if default is None:
        default = []
    if not os.path.exists(file_path):
        return default
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, Exception):
        return default


def save_json_atomic(file_path, data):
    """Public atomic JSON write."""
    _atomic_write_json(file_path, data)


# =====================================================================
# JSON TRACKER (sent_summaries.json)
# =====================================================================
def load_all_entries_from_json():
    """Return full list of entries from sent_summaries.json."""
    return load_json_safe(JSON_TRACKER_FILE, [])


def load_processed_files_from_json():
    """Return set of file_names that have been processed."""
    entries = load_all_entries_from_json()
    return {e["file_name"] for e in entries if e.get("file_name")}


def build_entry_from_metadata(file_name, metadata, parsing_notes="Success", subtopic=None):
    """Build a sent_summaries.json entry dict from extraction metadata."""
    system = ", ".join(metadata.get("specialty", [])) if isinstance(metadata.get("specialty"), list) else metadata.get("system", "Other")
    if subtopic is None:
        subtopic = system
    return {
        "serial_number": 0,
        "file_name": file_name,
        "title": metadata.get("title", metadata.get("paper_name", "Unknown Title")),
        "authors": metadata.get("authors", metadata.get("primary_authors", "Unknown Authors")),
        "journal": metadata.get("journal", metadata.get("journal_name", "Unknown Journal")),
        "doi": metadata.get("doi", "None"),
        "year": metadata.get("year", ""),
        "system": system,
        "type": metadata.get("article_subtype", metadata.get("doc_type", metadata.get("type_of_article", "Other"))),
        "subtopic": subtopic,
        "md_generated": "Yes",
        "email_pushed": "No",
        "date_added": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "email_pushed_date": "",
        "parsing_notes": parsing_notes,
        "show_on_web": "No",
    }


def log_transaction_to_json(file_name, metadata, parsing_notes="Success", subtopic=None):
    """Append one entry to sent_summaries.json atomically."""
    entries = load_all_entries_from_json()
    next_serial = len(entries) + 1
    entry = build_entry_from_metadata(file_name, metadata, parsing_notes, subtopic=subtopic)
    entry["serial_number"] = next_serial
    entries.append(entry)
    _atomic_write_json(JSON_TRACKER_FILE, entries)


# =====================================================================
# EXCEL TRACKER (sent_summaries.xlsx)
# =====================================================================
def initialize_excel_tracker():
    """Create Excel tracker if missing, migrate schema if needed."""
    if not os.path.exists(EXCEL_TRACKER_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registry Logs"
        ws.append(EXCEL_HEADERS)
        wb.save(EXCEL_TRACKER_FILE)
        print(f"  [OK] Excel tracker initialized: {EXCEL_TRACKER_FILE}")
    else:
        _migrate_ledger_schema()


def _migrate_ledger_schema():
    """Add missing columns to existing Excel ledger."""
    try:
        wb = load_workbook(EXCEL_TRACKER_FILE)
        ws = wb["Registry Logs"]
        headers = [cell.value for cell in ws[1]]
        if "Year" not in headers:
            col_idx = len(headers) + 1
            ws.cell(row=1, column=col_idx, value="Year")
            wb.save(EXCEL_TRACKER_FILE)
            print(f"  Migrated ledger: added Year column")
    except Exception:
        pass


def log_transaction_to_excel(file_name, metadata, parsing_notes="Success"):
    """Append one row to the Excel tracker (with retry on lock)."""
    retries = 5
    for attempt in range(retries):
        try:
            wb = load_workbook(EXCEL_TRACKER_FILE)
            ws = wb["Registry Logs"]
            next_serial = ws.max_row
            row_data = [
                next_serial, file_name,
                metadata.get("title", metadata.get("paper_name", "Unknown Title")),
                metadata.get("authors", metadata.get("primary_authors", "Unknown Authors")),
                metadata.get("journal", metadata.get("journal_name", "Unknown Journal")),
                metadata.get("doi", "None"),
                metadata.get("year", ""),
                ", ".join(metadata.get("specialty", [])) if isinstance(metadata.get("specialty"), list) else metadata.get("system", "Other"),
                metadata.get("article_subtype", metadata.get("doc_type", metadata.get("type_of_article", "Other"))),
                "Yes", "No",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "",
                parsing_notes, "No"
            ]
            ws.append(row_data)
            wb.save(EXCEL_TRACKER_FILE)
            return
        except PermissionError:
            time.sleep(1.5)


# =====================================================================
# REMOVED TRACKER
# =====================================================================
def load_removed_entries():
    """Load removed entries from sent_summaries_removed.json."""
    return load_json_safe(REMOVED_TRACKER_FILE, [])


def append_removed_entries(entries):
    """Append removed entries with timestamp."""
    removed = load_removed_entries()
    for entry in entries:
        entry["removed_date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        removed.append(entry)
    _atomic_write_json(REMOVED_TRACKER_FILE, removed)


# =====================================================================
# PEARLS TRACKER
# =====================================================================
def load_pearl_tracker(tracker_file):
    """Load set of already-processed file_names from pearl tracker."""
    if not os.path.exists(tracker_file):
        return set()
    try:
        wb = load_workbook(tracker_file, read_only=True)
        ws = wb["Tracker"]
        processed = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                processed.add(str(row[0]).strip())
        return processed
    except Exception:
        return set()


def update_pearl_tracker(tracker_file, file_name, pearl_count, mode):
    """Record a processed file in the pearl tracker."""
    try:
        if os.path.exists(tracker_file):
            wb = load_workbook(tracker_file)
            ws = wb["Tracker"]
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Tracker"
            ws.append(["file_name", "timestamp_processed", "pearl_count", "mode"])
        ws.append([file_name, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), pearl_count, mode])
        wb.save(tracker_file)
    except Exception as e:
        print(f"  Warning: failed to update pearl tracker: {e}")


# =====================================================================
# SUBTOPIC TRACKING - pending queue + completed mappings
# =====================================================================
def load_pending_subtopics():
    """Load pending_subtopics.json. Returns list."""
    return load_json_safe(getattr(__import__('acumen_core.config', fromlist=['PENDING_SUBTOPICS_FILE']), 'PENDING_SUBTOPICS_FILE', 'pending_subtopics.json'), [])


def save_pending_subtopics(entries):
    """Save pending_subtopics.json atomically."""
    from acumen_core.config import PENDING_SUBTOPICS_FILE
    _atomic_write_json(PENDING_SUBTOPICS_FILE, entries)


def append_pending_subtopic(title, system, type_val, file_name, timestamp=None):
    """Add one entry to pending_subtopics.json."""
    if timestamp is None:
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    entries = load_pending_subtopics()
    entries.append({
        "title": title,
        "system": system,
        "type": type_val,
        "file_name": file_name,
        "subtopic": system,
        "timestamp": timestamp,
        "processed": False,
    })
    save_pending_subtopics(entries)
    return len(entries)


def load_subtopic_mapping():
    """Load subtopic_mapping.json. Returns list."""
    from acumen_core.config import SUBTOPIC_MAPPING_FILE
    return load_json_safe(SUBTOPIC_MAPPING_FILE, [])


def save_subtopic_mapping(entries):
    """Save subtopic_mapping.json atomically."""
    from acumen_core.config import SUBTOPIC_MAPPING_FILE
    _atomic_write_json(SUBTOPIC_MAPPING_FILE, entries)


def append_subtopic_mapping(title, system, type_val, file_name, subtopic, timestamp=None):
    """Add one entry to subtopic_mapping.json (cumulative registry)."""
    if timestamp is None:
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    entries = load_subtopic_mapping()
    entries.append({
        "title": title,
        "system": system,
        "type": type_val,
        "file_name": file_name,
        "subtopic": subtopic,
        "timestamp": timestamp,
        "processed": True,
    })
    save_subtopic_mapping(entries)
    return len(entries)
